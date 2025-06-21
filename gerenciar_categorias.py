import pandas as pd
import openpyxl
from openpyxl import load_workbook

def adicionar_item_categoria(categoria, novo_item):
    """
    Adiciona um novo item a uma categoria específica na planilha
    
    Args:
        categoria (str): Nome da categoria (coluna)
        novo_item (str): Novo item a ser adicionado
    """
    try:
        # Carregar a planilha
        wb = load_workbook('Base_financas.xlsx')
        
        # Acessar a aba 'Despesas Categoria'
        ws = wb['Despesas Categoria']
        
        # Encontrar a coluna da categoria
        categoria_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == categoria:
                categoria_col = col
                break
        
        if categoria_col is None:
            print(f"❌ Categoria '{categoria}' não encontrada!")
            return False
        
        # Encontrar a primeira célula vazia na coluna
        row = 2  # Começar da linha 2 (após o cabeçalho)
        while ws.cell(row=row, column=categoria_col).value is not None:
            row += 1
        
        # Adicionar o novo item
        ws.cell(row=row, column=categoria_col, value=novo_item)
        
        # Salvar a planilha
        wb.save('Base_financas.xlsx')
        wb.close()
        
        print(f"✅ Item '{novo_item}' adicionado à categoria '{categoria}' com sucesso!")
        return True
        
    except Exception as e:
        print(f"❌ Erro ao adicionar item: {e}")
        return False

def listar_categorias():
    """Lista todas as categorias disponíveis"""
    try:
        df = pd.read_excel('Base_financas.xlsx', sheet_name='Despesas Categoria')
        print("📋 Categorias disponíveis:")
        for i, categoria in enumerate(df.columns, 1):
            print(f"{i}. {categoria}")
        return df.columns.tolist()
    except Exception as e:
        print(f"❌ Erro ao listar categorias: {e}")
        return []

def listar_itens_categoria(categoria):
    """Lista todos os itens de uma categoria específica"""
    try:
        df = pd.read_excel('Base_financas.xlsx', sheet_name='Despesas Categoria')
        if categoria in df.columns:
            itens = df[categoria].dropna().tolist()
            print(f"📋 Itens da categoria '{categoria}':")
            for i, item in enumerate(itens, 1):
                print(f"{i}. {item}")
            return itens
        else:
            print(f"❌ Categoria '{categoria}' não encontrada!")
            return []
    except Exception as e:
        print(f"❌ Erro ao listar itens: {e}")
        return []

if __name__ == "__main__":
    print("🔧 Gerenciador de Categorias e Itens")
    print("=" * 40)
    
    while True:
        print("\nEscolha uma opção:")
        print("1. Listar categorias")
        print("2. Listar itens de uma categoria")
        print("3. Adicionar novo item a uma categoria")
        print("4. Sair")
        
        opcao = input("\nDigite sua opção (1-4): ").strip()
        
        if opcao == "1":
            listar_categorias()
            
        elif opcao == "2":
            categorias = listar_categorias()
            if categorias:
                categoria = input("Digite o nome da categoria: ").strip()
                listar_itens_categoria(categoria)
                
        elif opcao == "3":
            categorias = listar_categorias()
            if categorias:
                categoria = input("Digite o nome da categoria: ").strip()
                if categoria in categorias:
                    novo_item = input("Digite o novo item: ").strip()
                    if novo_item:
                        adicionar_item_categoria(categoria, novo_item)
                    else:
                        print("❌ Nome do item não pode estar vazio!")
                else:
                    print("❌ Categoria não encontrada!")
                    
        elif opcao == "4":
            print("👋 Até logo!")
            break
            
        else:
            print("❌ Opção inválida! Digite 1, 2, 3 ou 4.") 