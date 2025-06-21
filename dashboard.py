import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image
from datetime import datetime, timedelta
import pytz
import os
import streamlit.components.v1 as components
from numerize import numerize
from streamlit_option_menu import option_menu
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
import openpyxl

# --- IMPORTAÇÕES DOS NOVOS SISTEMAS ---
from backup_system import BackupSystem, safe_backup
from crud_system import CRUDSystem, format_dataframe_for_display, create_editable_table

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="Dashboard Financeiro",
    page_icon="💲",
    layout="wide",
)

try: # --- BLOCO DE CAPTURA DE ERRO GLOBAL ---

    # --- FUNÇÃO PARA VERIFICAR E ADICIONAR COLUNA ---
    def check_and_add_favorecido_column(excel_path):
        """Verifica se a coluna 'FAVORECIDO' existe na aba 'Receitas' e a adiciona se não existir."""
        try:
            workbook = openpyxl.load_workbook(excel_path)
            if 'Receitas' in workbook.sheetnames:
                sheet = workbook['Receitas']
                headers = [cell.value for cell in sheet[1]]
                if 'FAVORECIDO' not in headers:
                    new_col_idx = sheet.max_column + 1
                    sheet.cell(row=1, column=new_col_idx, value='FAVORECIDO')
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row=row, column=new_col_idx, value='N/A')
                    workbook.save(excel_path)
                    st.toast("Coluna 'FAVORECIDO' adicionada com sucesso à aba 'Receitas'!")
            workbook.close()
        except Exception as e:
            st.warning(f"Atenção: Não foi possível adicionar a coluna 'FAVORECIDO' automaticamente: {e}")

    # --- CAMINHO DO ARQUIVO E CARREGAMENTO INICIAL ---
    excel_path = 'Base_financas.xlsx'
    check_and_add_favorecido_column(excel_path) # Executa a verificação no início
    xls = pd.ExcelFile(excel_path)

    # Outras importações úteis
    import plotly.graph_objects as go
    from datetime import datetime, timedelta

    # Importação do módulo de relatórios
    from relatorio_pdf import RelatorioPDF, create_download_button

    # Função para carregar itens de validação das categorias
    def carregar_itens_categoria():
        """Carrega os itens de cada categoria da planilha para validação"""
        try:
            xls = pd.ExcelFile(excel_path)
            
            # Carregar itens de despesas
            df_cat_desp = pd.read_excel(xls, sheet_name='Despesas Categoria')
            itens_despesas = {}
            for coluna in df_cat_desp.columns:
                itens = df_cat_desp[coluna].dropna().tolist()
                itens_despesas[coluna] = itens
            
            # Carregar itens de receitas
            df_cat_rec = pd.read_excel(xls, sheet_name='Receitas Categoria')
            itens_receitas = []
            if 'SUBCATEGORIA' in df_cat_rec.columns:
                itens_receitas = df_cat_rec['SUBCATEGORIA'].dropna().unique().tolist()
            
            return itens_despesas, itens_receitas
        except Exception as e:
            st.error(f"Erro ao carregar itens de categoria: {e}")
            return {}, []

    # Inicialização do session_state para os formulários
    if "show_despesa_form" not in st.session_state:
        st.session_state.show_despesa_form = False
    if "show_receita_form" not in st.session_state:
        st.session_state.show_receita_form = False
    if "show_cc_form" not in st.session_state:
        st.session_state.show_cc_form = False
    if "show_venda_form" not in st.session_state:
        st.session_state.show_venda_form = False

    # --- INICIALIZAÇÃO DOS SISTEMAS CRUD E BACKUP ---
    crud_system = CRUDSystem(excel_path)
    backup_system = BackupSystem(excel_path)
    
    # Inicialização do session_state para CRUD
    sheets_with_crud = ['Despesas', 'Receitas', 'Vendas', 'Investimentos', 'Div_CC']
    for sheet in sheets_with_crud:
        if f"show_edit_{sheet}" not in st.session_state:
            st.session_state[f"show_edit_{sheet}"] = False
        if f"show_delete_{sheet}" not in st.session_state:
            st.session_state[f"show_delete_{sheet}"] = False
        if f"show_bulk_delete_{sheet}" not in st.session_state:
            st.session_state[f"show_bulk_delete_{sheet}"] = False
        if f"selected_row_{sheet}" not in st.session_state:
            st.session_state[f"selected_row_{sheet}"] = None
        if f"selected_rows_{sheet}" not in st.session_state:
            st.session_state[f"selected_rows_{sheet}"] = []

    st.set_page_config(page_title="DashBoard", layout="wide")

    st.markdown("""
    <style>
        /* Aumenta a largura da sidebar para evitar quebra de texto nos filtros */
        section[data-testid="stSidebar"] {
            width: 350px !important;
        }
        
        /* Estilos para os novos cards de métricas */
        .metric-card {
            background-color: #1c213c;
            border: 1px solid #293153;
            border-radius: 12px;
            padding: 25px 20px;
            margin-bottom: 10px;
            color: #f0f2f6;
            position: relative;
        }
        
        /* Corrige a quebra de linha no menu de navegação */
        div[data-testid="stOptionMenu"] > ul {
            white-space: nowrap;
        }
        div[data-testid="stOptionMenu"] > ul > li {
            display: inline-block;
            padding: 0 10px; /* Adiciona um espaçamento mais uniforme */
        }
        
        .metric-card.saldo {
            background: linear-gradient(to right, #e96443, #904e95);
            border: none;
        }
        .metric-card .title-row {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 8px;
        }
        .metric-card h4 {
            font-size: 1rem;
            font-weight: 400;
            margin: 0;
            opacity: 0.8;
            color: #d1d5db;
        }
        .metric-card.saldo h4, .metric-card.saldo h2 {
            color: white;
        }
        .metric-card h2 {
            font-size: 2.2rem;
            font-weight: 700;
            margin: 0;
            color: #f9fafb;
        }
        .metric-card .delta {
            font-size: 0.8rem;
            font-weight: 500;
            background-color: rgba(255, 255, 255, 0.1);
            padding: 2px 8px;
            border-radius: 8px;
        }
        .delta-p-green { color: #34d399; }
        .delta-p-red { color: #f87171; }

        /* Correção para o menu de navegação (option_menu) */
        ul[class^="nav"] {
            display: flex;
            flex-wrap: nowrap !important; /* Impede a quebra de linha */
            white-space: nowrap;
            overflow-x: auto; /* Adiciona scroll horizontal se necessário */
            -webkit-overflow-scrolling: touch; /* Melhora a rolagem em dispositivos móveis */
        }

        li[class^="nav-item"] {
            flex-shrink: 0; /* Impede que os itens do menu encolham */
        }

        a[class^="nav-link"] {
            display: flex !important;
            align-items: center !important; /* Alinha ícone e texto verticalmente */
            white-space: nowrap !important; /* Garante que o texto dentro do link não quebre */
            padding: 10px 15px !important; /* Ajusta o espaçamento interno */
        }

        a[class^="nav-link"] > i {
            margin-right: 8px !important; /* Adiciona espaço entre o ícone e o texto */
            padding-bottom: 2px; /* Ajuste fino do alinhamento vertical do ícone */
        }

    </style>
    """, unsafe_allow_html=True)

    # Ícone de finanças (exemplo: 💰)
    st.markdown("## 💰 Análise Descritiva de Finanças Pessoais")

    # --- SISTEMA DE NAVEGAÇÃO ---
    selected = option_menu(
        menu_title=None,
        options=["Visão Geral", "Transações", "Para onde vai", "Pra quem vai", "Classificação ABC", "Fluxo de Caixa", "Despesas por Categoria", "Vendas", "Investimentos", "Cartão de Crédito", "Orçamento"],
        icons=["bar-chart", "table", "graph-up", "people-fill", "list-ol", "graph-up-arrow", "bar-chart-fill", "cart-fill", "piggy-bank", "credit-card", "target"],
        orientation="horizontal",
        default_index=0,
    )

    st.header(f"Aba Selecionada: '{selected}'") # DEBUG PARA VER O VALOR DA ABA

    # Lê todas as abas do arquivo Excel
    abas = xls.sheet_names

    # Exemplo de leitura de uma aba específica:
    # df_despesas = pd.read_excel(xls, sheet_name='Despesas')

    # --- BARRA LATERAL (SIDEBAR) ---
    with st.sidebar:
        st.image("https://raw.githubusercontent.com/Nandho1/Nandho1/main/Nandho1.gif", width=250)
        
        # Obter data atual para filtros padrão
        hoje = datetime.now(pytz.timezone('America/Sao_Paulo'))
        ano_atual = str(hoje.year)
        mes_ingles = hoje.strftime('%b').capitalize()
        # Mapeamento para traduzir os meses
        mapa_meses = {'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'}
        mes_atual = mapa_meses.get(mes_ingles, mes_ingles)

        # Filtros
        st.header("Filtros")
        
        # Carregar anos de todas as fontes de dados, garantindo a conversão para datetime
        df_r = pd.read_excel(xls, sheet_name='Receitas')
        df_d = pd.read_excel(xls, sheet_name='Despesas')
        df_i = pd.read_excel(xls, sheet_name='Investimentos')
        df_c = pd.read_excel(xls, sheet_name='Div_CC')
        df_v = pd.read_excel(xls, sheet_name='Vendas')

        anos_r = pd.to_datetime(df_r['DATA'], errors='coerce').dt.year
        anos_d = pd.to_datetime(df_d['DATA'], errors='coerce').dt.year
        anos_i = pd.to_datetime(df_i['DATA'], errors='coerce').dt.year
        anos_c = pd.to_datetime(df_c['Data'], errors='coerce').dt.year # Cuidado com 'Data' vs 'DATA'
        anos_v = pd.to_datetime(df_v['DATA'], errors='coerce').dt.year

        anos_disponiveis = sorted(pd.concat([anos_r, anos_d, anos_i, anos_c, anos_v]).dropna().unique().astype(int).astype(str))
        
        # Definir ano_atual como padrão se estiver na lista, senão o último disponível
        default_ano = [ano_atual] if ano_atual in anos_disponiveis else [anos_disponiveis[-1]] if anos_disponiveis else []
        anos_selecionados = st.multiselect('Ano', anos_disponiveis, default=default_ano)

        meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        # Definir mes_atual como padrão se houver anos selecionados
        default_mes = [mes_atual] if anos_selecionados else []
        meses_selecionados = st.multiselect('Mês', meses_ordem, default=default_mes)
        
        # Contas
        df_conta = pd.read_excel(xls, sheet_name='Conta')
        contas = df_conta['Contas'].dropna().unique().tolist()

        # Filtro de tipo de análise com checkboxes para evitar texto cortado
        st.write('**Tipo de análise**')
        col_rec, col_desp = st.columns(2)
        show_receitas = col_rec.checkbox('Receitas', value=True)
        show_despesas = col_desp.checkbox('Despesas', value=True)

        tipos_selecionados = []
        if show_receitas:
            tipos_selecionados.append('Receitas')
        if show_despesas:
            tipos_selecionados.append('Despesas')

        # Filtro de conta com opção 'Todas'
        contas_opcoes = ['Todas'] + contas
        if 'reset_filtros' not in st.session_state:
            st.session_state['reset_filtros'] = False
            
        if st.session_state['reset_filtros']:
            conta_selecionada = st.selectbox('Conta', contas_opcoes, index=0, key='conta_filtro_reset')
        else:
            conta_selecionada = st.selectbox('Conta', contas_opcoes, key='conta_filtro')

        # Filtro dinâmico de categoria
        if 'Receitas' in tipos_selecionados:
            df_cat = pd.read_excel(xls, sheet_name='Receitas Categoria')
            categorias = df_cat['SUBCATEGORIA'].dropna().unique().tolist()
        else:
            df_cat = pd.read_excel(xls, sheet_name='Despesas Categoria')
            categorias = df_cat.columns.tolist()

        # Filtro de categoria com opção 'Todas'
        categorias_opcoes = ['Todas'] + categorias
        if st.session_state['reset_filtros']:
            categoria_selecionada = st.selectbox('Categoria', categorias_opcoes, index=0, key='cat_filtro_reset')
        else:
            categoria_selecionada = st.selectbox('Categoria', categorias_opcoes, key='cat_filtro')

        st.markdown('---')
        if st.button('Redefinir Filtros'):
            st.session_state['reset_filtros'] = True
        
        # Filtros específicos por aba (movidos para dentro do sidebar)
        # Só mostra o expander de investimentos se a aba Investimentos estiver selecionada
        if selected == 'Investimentos':
            df_investimentos = pd.read_excel(xls, sheet_name='Investimentos')
            df_investimentos['DATA'] = pd.to_datetime(df_investimentos['DATA'], errors='coerce')
            df_investimentos = df_investimentos.dropna(subset=['DATA'])
            df_investimentos['Ano'] = df_investimentos['DATA'].dt.year.astype(str)
            df_investimentos['Mês'] = df_investimentos['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            with st.expander('Filtros de Investimentos', expanded=True):
                anos_invest = df_investimentos['Ano'].dropna().unique().tolist()
                anos_invest_sel = st.multiselect('Ano (Investimentos)', sorted(anos_invest, reverse=True), default=sorted(anos_invest, reverse=True), key='ano_invest')
                meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                meses_invest = [m for m in meses_ordem if m in df_investimentos['Mês'].unique()]
                meses_invest_sel = st.multiselect('Mês (Investimentos)', meses_invest, default=meses_invest, key='mes_invest')
                tipos_invest = ['Todos'] + df_investimentos['TIPO'].dropna().unique().tolist()
                tipo_invest_sel = st.selectbox('Tipo de Investimento', tipos_invest, key='tipo_invest')
                objetivos_invest = ['Todos'] + df_investimentos['OBJETIVO'].dropna().unique().tolist()
                objetivo_invest_sel = st.selectbox('Objetivo', objetivos_invest, key='objetivo_invest')
                ativos_invest = ['Todos'] + df_investimentos['ATIVO'].dropna().unique().tolist()
                ativo_invest_sel = st.selectbox('Ativo', ativos_invest, key='ativo_invest')
        
        # Só mostra o expander de Cartão de Crédito se a aba estiver selecionada
        if selected == 'Cartão de Crédito':
            df_cc = pd.read_excel(xls, sheet_name='Div_CC')
            df_cc['Data'] = pd.to_datetime(df_cc['Data'], errors='coerce')
            df_cc.dropna(subset=['Data'], inplace=True)
            df_cc['Ano'] = df_cc['Data'].dt.year.astype(str)
            df_cc['Mês'] = df_cc['Data'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            with st.expander('Filtros do Cartão de Crédito', expanded=True):
                anos_cc = df_cc['Ano'].dropna().unique().tolist()
                anos_cc_sel = st.multiselect('Ano (Cartão)', sorted(anos_cc, reverse=True), default=sorted(anos_cc, reverse=True), key='ano_cc_sidebar')
                
                meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                meses_cc = [m for m in meses_ordem if m in df_cc['Mês'].unique()]
                meses_cc_sel = st.multiselect('Mês (Cartão)', meses_cc, default=meses_cc, key='mes_cc_sidebar')

                cartoes_cc = ['Todos'] + df_cc['Cartão'].dropna().unique().tolist()
                cartao_cc_sel = st.selectbox('Cartão', cartoes_cc, key='cartao_cc_sidebar')

                situacao_cc = ['Todas'] + df_cc['Situação'].dropna().unique().tolist()
                situacao_cc_sel = st.selectbox('Situação', situacao_cc, key='situacao_cc_sidebar')

                tipos_compra = ['Todos'] + df_cc['Tipo de Compra'].dropna().unique().tolist()
                tipo_compra_sel = st.selectbox('Tipo de Compra', tipos_compra, key='tipo_compra_cc_sidebar')

                parcelas_opcoes = ['Todas'] + [str(x) for x in sorted(df_cc['Quantidade de parcelas'].dropna().unique())]
                parcelas_sel = st.selectbox('Quantidade de Parcelas', parcelas_opcoes, key='parcelas_cc_sidebar')

        # --- SEÇÃO DE GERENCIAMENTO DE BACKUP ---
        st.markdown('---')
        with st.expander("💾 Gerenciamento de Backup", expanded=False):
            col_backup1, col_backup2 = st.columns(2)
            
            with col_backup1:
                if st.button("🔄 Criar Backup", use_container_width=True):
                    success, message = safe_backup("manual")
                    if success:
                        st.success("Backup criado com sucesso!")
                    else:
                        st.error(f"Erro ao criar backup: {message}")
            
            with col_backup2:
                if st.button("📋 Listar Backups", use_container_width=True):
                    backups = backup_system.list_backups()
                    if backups:
                        st.write("**Backups disponíveis:**")
                        for backup in backups[:5]:  # Mostra apenas os 5 mais recentes
                            st.write(f"📁 {backup['filename']} ({backup['date'].strftime('%d/%m/%Y %H:%M')})")
                    else:
                        st.info("Nenhum backup encontrado.")
            
            # Opção para restaurar backup
            backups = backup_system.list_backups()
            if backups:
                backup_options = [f"{b['filename']} ({b['date'].strftime('%d/%m/%Y %H:%M')})" for b in backups[:5]]
                selected_backup = st.selectbox("Selecionar backup para restaurar:", backup_options, key="backup_restore")
                
                if st.button("🔄 Restaurar Backup", use_container_width=True):
                    if selected_backup:
                        backup_filename = selected_backup.split(" (")[0]
                        backup_path = os.path.join('backups', backup_filename)
                        success, message = backup_system.restore_backup(backup_path)
                        if success:
                            st.success("Backup restaurado com sucesso! Recarregue a página.")
                            st.rerun()
                        else:
                            st.error(f"Erro ao restaurar backup: {message}")

    # --- CÁLCULO DOS INDICADORES BÁSICOS ---

    # Filtra receitas
    receitas = pd.read_excel(xls, sheet_name='Receitas')
    receitas['DATA'] = pd.to_datetime(receitas['DATA'], errors='coerce')
    receitas = receitas.dropna(subset=['DATA'])
    receitas['Ano'] = receitas['DATA'].dt.year.astype(str)
    receitas['Mês'] = receitas['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})

    receitas_filtradas = receitas[
        ((receitas['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas')) &
        (receitas['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
        (receitas['Mês'].isin(meses_selecionados) if meses_selecionados else True)
    ]

    # Categoria de receitas
    if 'Receitas' in tipos_selecionados and categoria_selecionada != 'Todas':
        receitas_filtradas = receitas_filtradas[receitas_filtradas['CATEGORIA'] == categoria_selecionada]

    # Filtra despesas

    despesas = pd.read_excel(xls, sheet_name='Despesas')
    despesas['DATA'] = pd.to_datetime(despesas['DATA'], errors='coerce')
    despesas = despesas.dropna(subset=['DATA'])
    despesas['Ano'] = despesas['DATA'].dt.year.astype(str)
    despesas['Mês'] = despesas['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})

    despesas_filtradas = despesas[
        ((despesas['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas')) &
        (despesas['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
        (despesas['Mês'].isin(meses_selecionados) if meses_selecionados else True)
    ]

    # Categoria de despesas
    if 'Despesas' in tipos_selecionados and categoria_selecionada != 'Todas':
        despesas_filtradas = despesas_filtradas[despesas_filtradas['CATEGORIA'] == categoria_selecionada]

    # Se o tipo não estiver selecionado, zera os valores
    if 'Receitas' not in tipos_selecionados:
        valor_recebidos = 0
    else:
        valor_recebidos = receitas_filtradas['VALOR'].sum()

    if 'Despesas' not in tipos_selecionados:
        valor_despesas = 0
    else:
        valor_despesas = despesas_filtradas['VALOR'].sum()

    saldo = valor_recebidos + valor_despesas  # Despesas já são negativas
    # Corrige o cálculo do percentual para evitar divisão por zero
    if valor_recebidos > 0:
        percentual = (abs(valor_despesas) / valor_recebidos * 100)
    else:
        percentual = 0.0

    # --- CÁLCULO DE PERÍODO ANTERIOR E DELTAS ---
    saldo_anterior = 0
    # Define um valor padrão caso o cálculo não seja possível
    delta_receitas = 0
    delta_despesas = 0

    # A comparação só faz sentido se um único mês e ano forem selecionados
    if len(meses_selecionados) == 1 and len(anos_selecionados) == 1:
        ano_atual_num = int(anos_selecionados[0])
        mes_atual_idx = meses_ordem.index(meses_selecionados[0])
        
        if mes_atual_idx > 0:
            mes_ant_str = meses_ordem[mes_atual_idx - 1]
            ano_ant_num = ano_atual_num
        else:
            # Janeiro -> Dezembro do ano anterior
            mes_ant_str = 'Dez'
            ano_ant_num = ano_atual_num - 1

        ano_ant_str = str(ano_ant_num)

        # Filtra dados do período anterior
        receitas_ant_df = receitas[
            (receitas['Ano'] == ano_ant_str) & (receitas['Mês'] == mes_ant_str)
        ]
        despesas_ant_df = despesas[
            (despesas['Ano'] == ano_ant_str) & (despesas['Mês'] == mes_ant_str)
        ]
        
        if conta_selecionada != 'Todas':
            receitas_ant_df = receitas_ant_df[receitas_ant_df['CONTA'] == conta_selecionada]
            despesas_ant_df = despesas_ant_df[despesas_ant_df['CONTA'] == conta_selecionada]

        receitas_anterior = receitas_ant_df['VALOR'].sum()
        despesas_anterior = despesas_ant_df['VALOR'].sum()
        saldo_anterior = receitas_anterior + despesas_anterior

        # Calcular deltas
        if receitas_anterior > 0:
            delta_receitas = ((valor_recebidos - receitas_anterior) / receitas_anterior) * 100
        else:
            delta_receitas = 0.0
            
        if abs(despesas_anterior) > 0:
            delta_despesas = ((abs(valor_despesas) - abs(despesas_anterior)) / abs(despesas_anterior)) * 100
        else:
            delta_despesas = 0.0

    def format_brl(valor):
        # Verifica se o valor é NaN ou None
        if pd.isna(valor) or valor is None:
            return "R$ 0,00"
        
        # Converte para float se necessário
        try:
            valor = float(valor)
        except (ValueError, TypeError):
            return "R$ 0,00"
        
        return f"R$ {valor:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")

    # Helper para HTML do delta
    def get_delta_html(delta, reverse_colors=False):
        # Verifica se o delta é NaN ou None
        if pd.isna(delta) or delta is None:
            return ""
        
        # Converte para float se necessário
        try:
            delta = float(delta)
        except (ValueError, TypeError):
            return ""
        
        if delta > 0:
            arrow = "↑"
            color_class = "delta-p-green" if not reverse_colors else "delta-p-red"
        elif delta < 0:
            arrow = "↓"
            color_class = "delta-p-red" if not reverse_colors else "delta-p-green"
        else:
            arrow = ""
            color_class = ""
        
        return f'<span class="delta {color_class}">{abs(delta):.1f}% {arrow}</span>' if delta != 0 else ''

    # Exibe cards principais apenas se não estiver na aba Vendas
    if selected != 'Vendas' and selected != 'Cartão de Crédito' and selected != 'Investimentos':
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            st.markdown(f"""
            <div class="metric-card saldo">
                <h4>Saldo</h4>
                <h2>{format_brl(saldo)}</h2>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            delta_html = get_delta_html(delta_receitas)
            st.markdown(f"""
            <div class="metric-card">
                <div class="title-row">
                    <h4>Recebimentos</h4>{delta_html}
                </div>
                <h2>{format_brl(valor_recebidos)}</h2>
            </div>
            """, unsafe_allow_html=True)

        with col3:
            delta_html = get_delta_html(delta_despesas, reverse_colors=True)
            st.markdown(f"""
            <div class="metric-card">
                <div class="title-row">
                    <h4>Despesas</h4>{delta_html}
                </div>
                <h2>{format_brl(abs(valor_despesas))}</h2>
            </div>
            """, unsafe_allow_html=True)

        with col4:
            saldo_ant_text = format_brl(saldo_anterior) if (len(meses_selecionados) == 1 and len(anos_selecionados) == 1) else "N/A"
            st.markdown(f"""
            <div class="metric-card">
                <h4>Saldo Anterior</h4>
                <h2>{saldo_ant_text}</h2>
            </div>
            """, unsafe_allow_html=True)

        with col5:
            st.markdown(f"""
            <div class="metric-card">
                <h4>Despesas vs Receitas %</h4>
                <h2>{percentual:.1f}%</h2>
            </div>
            """, unsafe_allow_html=True)

    if selected == "Visão Geral":
        # Botão de exportação de relatório
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("📄 Exportar Relatório PDF", type="primary", use_container_width=True):
                # Criar instância do gerador de relatórios
                relatorio = RelatorioPDF()
                
                # Preparar dados para o relatório
                dados_relatorio = {
                    'saldo': saldo,
                    'receitas': valor_recebidos,
                    'despesas': abs(valor_despesas),
                    'percentual': percentual,
                    'graficos': []
                }
                
                # Adicionar gráficos se existirem dados
                if not despesas_filtradas.empty:
                    # Gráfico de evolução das despesas
                    df_evolucao = despesas_filtradas.copy()
                    df_evolucao['Dia'] = df_evolucao['DATA'].dt.day
                    evolucao = (
                        df_evolucao.groupby('Dia')['VALOR']
                        .sum()
                        .abs()
                        .reset_index()
                        .sort_values('Dia')
                    )
                    fig_evol = px.bar(
                        evolucao,
                        x='Dia',
                        y='VALOR',
                        labels={'Dia': 'Dia do Mês', 'VALOR': 'Despesas (R$)'},
                        title='Evolução das Despesas ao Longo dos Dias',
                    )
                    dados_relatorio['graficos'].append({
                        'titulo': 'Evolução das Despesas ao Longo dos Dias',
                        'fig': fig_evol,
                        'descricao': 'Análise da distribuição das despesas por dia do mês'
                    })
                
                # Gerar relatório
                try:
                    filename = relatorio.generate_dashboard_report(dados_relatorio, "relatorio_visao_geral.pdf")
                    st.success("✅ Relatório gerado com sucesso!")
                    create_download_button(filename, "📥 Download Relatório Visão Geral")
                except Exception as e:
                    st.error(f"❌ Erro ao gerar relatório: {e}")
        
        st.markdown("---")
        
        # --- GRÁFICOS DE ROSCA LADO A LADO ---
        cols = st.columns([1, 0.08, 1, 0.08, 1])
        col_g1, col_g2, col_g3 = cols[0], cols[2], cols[4]

        with col_g1:
            # Gráfico 1 (Categoria)
            df_top5 = despesas_filtradas.copy()
            if not df_top5.empty:
                top5 = (
                    df_top5.groupby('CATEGORIA')['VALOR']
                    .sum()
                    .abs()
                    .sort_values(ascending=False)
                    .head(5)
                )
                fig = px.pie(
                    names=top5.index,
                    values=top5.values,
                    hole=0.5,
                    title='Top 5 Despesas por Categoria',
                )
                fig.update_traces(
                    textinfo='percent',
                    textposition='outside',
                    pull=[0.05]*5,
                    textfont_size=16,
                )
                fig.update_layout(
                    legend=dict(orientation='h', yanchor='bottom', y=-0.35, xanchor='center', x=0.5, font=dict(size=11))
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info('Não há despesas para exibir no gráfico com os filtros selecionados.')

        with col_g2:
            # Gráfico 2 (Descrição)
            df_top5_desc = despesas_filtradas.copy()
            if not df_top5_desc.empty:
                top5_desc = (
                    df_top5_desc.groupby('DESCRIÇÃO')['VALOR']
                    .sum()
                    .abs()
                    .sort_values(ascending=False)
                    .head(5)
                )
                fig_desc = px.pie(
                    names=top5_desc.index,
                    values=top5_desc.values,
                    hole=0.5,
                    title='Top 5 Despesas por Descrição',
                )
                fig_desc.update_traces(
                    textinfo='percent',
                    textposition='outside',
                    pull=[0.05]*5,
                    textfont_size=16,
                )
                fig_desc.update_layout(
                    legend=dict(orientation='h', yanchor='bottom', y=-0.35, xanchor='center', x=0.5)
                )
                st.plotly_chart(fig_desc, use_container_width=True)
            else:
                st.info('Não há despesas para exibir no gráfico de descrição com os filtros selecionados.')

        with col_g3:
            # Gráfico 3 (Favorecido)
            df_top5_fav = despesas_filtradas.copy()
            if not df_top5_fav.empty:
                top5_fav = (
                    df_top5_fav.groupby('FAVORECIDO')['VALOR']
                    .sum()
                    .abs()
                    .sort_values(ascending=False)
                    .head(5)
                )
                fig_fav = px.pie(
                    names=top5_fav.index,
                    values=top5_fav.values,
                    hole=0.5,
                    title='Top 5 Despesas por Favorecido',
                )
                fig_fav.update_traces(
                    textinfo='percent',
                    textposition='outside',
                    pull=[0.05]*5,
                    textfont_size=16,
                )
                fig_fav.update_layout(
                    legend=dict(orientation='h', yanchor='bottom', y=-0.35, xanchor='center', x=0.5, font=dict(size=11))
                )
                st.plotly_chart(fig_fav, use_container_width=True)
            else:
                st.info('Não há despesas para exibir no gráfico de favorecido com os filtros selecionados.')

        # --- GRÁFICO DE EVOLUÇÃO DAS DESPESAS AO LONGO DOS DIAS ---
        st.markdown('---')
        df_evolucao = despesas_filtradas.copy()
        if not df_evolucao.empty:
            df_evolucao['Dia'] = df_evolucao['DATA'].dt.day
            evolucao = (
                df_evolucao.groupby('Dia')['VALOR']
                .sum()
                .abs()
                .reset_index()
                .sort_values('Dia')
            )
            fig_evol = px.bar(
                evolucao,
                x='Dia',
                y='VALOR',
                labels={'Dia': 'Dia do Mês', 'VALOR': 'Despesas (R$)'},
                title='Evolução das Despesas ao Longo dos Dias',
            )
            fig_evol.update_layout(
                xaxis=dict(dtick=1),
                yaxis_tickformat = ',.2f',
                yaxis_title='Despesas (R$)',
                xaxis_title='Dia do Mês',
                bargap=0.2
            )
            st.plotly_chart(fig_evol, use_container_width=True)
        else:
            st.info('Não há despesas para exibir a evolução ao longo dos dias com os filtros selecionados.')

    elif selected == "Transações":
        st.markdown('### Transações Detalhadas')
        
        # Carrega dados originais para CRUD
        df_despesas_crud = crud_system.load_sheet_data('Despesas')
        df_receitas_crud = crud_system.load_sheet_data('Receitas')
        
        # Aplica filtros aos dados CRUD
        if not df_despesas_crud.empty:
            df_despesas_crud['DATA'] = pd.to_datetime(df_despesas_crud['DATA'], errors='coerce')
            df_despesas_crud = df_despesas_crud.dropna(subset=['DATA'])
            df_despesas_crud['Ano'] = df_despesas_crud['DATA'].dt.year.astype(str)
            df_despesas_crud['Mês'] = df_despesas_crud['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            df_despesas_crud_filtrado = df_despesas_crud[
                ((df_despesas_crud['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas')) &
                (df_despesas_crud['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
                (df_despesas_crud['Mês'].isin(meses_selecionados) if meses_selecionados else True)
            ]
            
            if 'Despesas' in tipos_selecionados and categoria_selecionada != 'Todas':
                df_despesas_crud_filtrado = df_despesas_crud_filtrado[df_despesas_crud_filtrado['CATEGORIA'] == categoria_selecionada]
        else:
            df_despesas_crud_filtrado = pd.DataFrame()
        
        if not df_receitas_crud.empty:
            df_receitas_crud['DATA'] = pd.to_datetime(df_receitas_crud['DATA'], errors='coerce')
            df_receitas_crud = df_receitas_crud.dropna(subset=['DATA'])
            df_receitas_crud['Ano'] = df_receitas_crud['DATA'].dt.year.astype(str)
            df_receitas_crud['Mês'] = df_receitas_crud['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            df_receitas_crud_filtrado = df_receitas_crud[
                ((df_receitas_crud['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas')) &
                (df_receitas_crud['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
                (df_receitas_crud['Mês'].isin(meses_selecionados) if meses_selecionados else True)
            ]
            
            if 'Receitas' in tipos_selecionados and categoria_selecionada != 'Todas':
                df_receitas_crud_filtrado = df_receitas_crud_filtrado[df_receitas_crud_filtrado['CATEGORIA'] == categoria_selecionada]
        else:
            df_receitas_crud_filtrado = pd.DataFrame()
        
        # Combina dados para exibição
        df_trans_crud = pd.concat([
            df_receitas_crud_filtrado.assign(TIPO='Receita'),
            df_despesas_crud_filtrado.assign(TIPO='Despesa')
        ], ignore_index=True)
        
        if not df_trans_crud.empty:
            df_trans_crud = df_trans_crud.sort_values('DATA')
            df_trans_crud['DATA'] = pd.to_datetime(df_trans_crud['DATA']).dt.strftime('%d/%m/%Y')
            
            # Formata valores para exibição
            df_trans_crud['VALOR_FORMATADO'] = df_trans_crud['VALOR'].apply(format_brl)
            
            # Seleciona colunas para exibição
            colunas_exibicao = ['DATA', 'TIPO', 'DESCRIÇÃO', 'CATEGORIA', 'CONTA', 'VALOR_FORMATADO']
            if 'FAVORECIDO' in df_trans_crud.columns:
                colunas_exibicao.insert(2, 'FAVORECIDO')
            
            df_exibicao = df_trans_crud[colunas_exibicao].copy()
            df_exibicao = df_exibicao.rename(columns={
                'DATA': 'Data',
                'TIPO': 'Tipo',
                'DESCRIÇÃO': 'Descrição',
                'CATEGORIA': 'Categoria',
                'CONTA': 'Conta',
                'VALOR_FORMATADO': 'Valor',
                'FAVORECIDO': 'Favorecido'
            })
            
            # Exibe tabela com funcionalidades CRUD
            st.dataframe(
                df_exibicao,
                use_container_width=True,
                hide_index=True
            )
            
            # Botões de ação CRUD
            st.markdown('---')
            col_crud1, col_crud2, col_crud3, col_crud4, col_crud5, col_crud6 = st.columns([1, 1, 1, 1, 1, 1])
            
            with col_crud1:
                if st.button("➕ Nova Despesa", key="nova_despesa_transacao"):
                    st.session_state.show_despesa_form = True
            
            with col_crud2:
                if st.button("💰 Nova Receita", key="nova_receita_transacao"):
                    st.session_state.show_receita_form = True
            
            with col_crud3:
                if st.button("🛒 Nova Venda", key="nova_venda_transacao"):
                    st.session_state.show_venda_form = True
            
            with col_crud4:
                if st.button("✏️ Editar Transação", key="edit_transacao"):
                    st.session_state.show_edit_transacao = True
            
            with col_crud5:
                if st.button("🗑️ Excluir Transação", key="delete_transacao"):
                    st.session_state.show_delete_transacao = True
            
            with col_crud6:
                if st.button("🗑️ Exclusão em Lote", key="bulk_delete_transacao"):
                    st.session_state.show_bulk_delete_transacao = True
            
            # Formulário de edição de transação
            if st.session_state.get("show_edit_transacao", False):
                st.markdown('### ✏️ Editar Transação')
                
                # Seleção do tipo de transação
                tipo_transacao = st.selectbox("Tipo de Transação", ["Receita", "Despesa"], key="edit_tipo_transacao")
                
                if tipo_transacao == "Receita":
                    df_edit = df_receitas_crud_filtrado
                    sheet_name = "Receitas"
                else:
                    df_edit = df_despesas_crud_filtrado
                    sheet_name = "Despesas"
                
                if not df_edit.empty:
                    # Seleção da transação para editar
                    opcoes_transacao = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['DESCRIÇÃO']} - {format_brl(row['VALOR'])}" 
                                       for idx, row in df_edit.iterrows()]
                    transacao_selecionada = st.selectbox("Selecionar transação para editar:", opcoes_transacao, key="edit_transacao_select")
                    
                    if transacao_selecionada:
                        # Encontra o índice da transação selecionada
                        idx_selecionado = opcoes_transacao.index(transacao_selecionada)
                        row_to_edit = df_edit.iloc[idx_selecionado]
                        
                        with st.form("edit_transacao_form"):
                            st.write("**Editar dados da transação:**")
                            
                            col1, col2 = st.columns(2)
                            with col1:
                                nova_data = col1.date_input("Data", value=pd.to_datetime(row_to_edit['DATA']).date(), key="edit_data")
                                nova_descricao = col1.text_input("Descrição", value=row_to_edit['DESCRIÇÃO'], key="edit_descricao")
                            
                            with col2:
                                nova_categoria = col2.text_input("Categoria", value=row_to_edit['CATEGORIA'], key="edit_categoria")
                                novo_valor = col2.number_input("Valor", value=float(row_to_edit['VALOR']), format="%.2f", key="edit_valor")
                            
                            if 'FAVORECIDO' in row_to_edit.index:
                                novo_favorecido = st.text_input("Favorecido", value=row_to_edit['FAVORECIDO'] if pd.notna(row_to_edit['FAVORECIDO']) else "", key="edit_favorecido")
                            
                            col_submit, col_cancel = st.columns(2)
                            submitted = col_submit.form_submit_button("💾 Salvar Alterações", use_container_width=True, type="primary")
                            if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                                st.session_state.show_edit_transacao = False
                                st.rerun()
                            
                            if submitted:
                                # Atualiza os dados
                                updated_data = {
                                    'DATA': pd.to_datetime(nova_data),
                                    'DESCRIÇÃO': nova_descricao,
                                    'CATEGORIA': nova_categoria,
                                    'VALOR': novo_valor
                                }
                                
                                if 'FAVORECIDO' in row_to_edit.index:
                                    updated_data['FAVORECIDO'] = novo_favorecido
                                
                                # Encontra o índice original no dataframe completo
                                original_idx = df_edit.index[idx_selecionado]
                                success, message = crud_system.update_record(sheet_name, original_idx, updated_data)
                                
                                if success:
                                    st.success("Transação atualizada com sucesso!")
                                    st.session_state.show_edit_transacao = False
                                    st.rerun()
                                else:
                                    st.error(f"Erro ao atualizar transação: {message}")
                else:
                    st.info(f"Nenhuma {tipo_transacao.lower()} encontrada com os filtros selecionados.")
            
            # Formulário de exclusão de transação
            if st.session_state.get("show_delete_transacao", False):
                st.markdown('### 🗑️ Excluir Transação')
                
                tipo_transacao = st.selectbox("Tipo de Transação", ["Receita", "Despesa"], key="delete_tipo_transacao")
                
                if tipo_transacao == "Receita":
                    df_delete = df_receitas_crud_filtrado
                    sheet_name = "Receitas"
                else:
                    df_delete = df_despesas_crud_filtrado
                    sheet_name = "Despesas"
                
                if not df_delete.empty:
                    opcoes_transacao = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['DESCRIÇÃO']} - {format_brl(row['VALOR'])}" 
                                       for idx, row in df_delete.iterrows()]
                    transacao_selecionada = st.selectbox("Selecionar transação para excluir:", opcoes_transacao, key="delete_transacao_select")
                    
                    if st.button("🗑️ Confirmar Exclusão", key="confirm_delete_transacao"):
                        if transacao_selecionada:
                            idx_selecionado = opcoes_transacao.index(transacao_selecionada)
                            original_idx = df_delete.index[idx_selecionado]
                            success, message = crud_system.delete_record(sheet_name, original_idx)
                            
                            if success:
                                st.success("Transação excluída com sucesso!")
                                st.session_state.show_delete_transacao = False
                                st.rerun()
                            else:
                                st.error(f"Erro ao excluir transação: {message}")
                else:
                    st.info(f"Nenhuma {tipo_transacao.lower()} encontrada com os filtros selecionados.")
        
        # Formulário de exclusão em lote de transações
        if st.session_state.get("show_bulk_delete_transacao", False):
            st.markdown('### 🗑️ Exclusão em Lote - Transações')
            
            # Seleção do tipo de transação
            tipo_transacao = st.selectbox("Tipo de Transação", ["Receita", "Despesa"], key="bulk_delete_tipo_transacao")
            
            if tipo_transacao == "Receita":
                df_bulk_delete = df_receitas_crud_filtrado
                sheet_name = "Receitas"
            else:
                df_bulk_delete = df_despesas_crud_filtrado
                sheet_name = "Despesas"
            
            if not df_bulk_delete.empty:
                # Criar opções para seleção múltipla
                opcoes_transacao = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['DESCRIÇÃO']} - {format_brl(row['VALOR'])}" 
                                   for idx, row in df_bulk_delete.iterrows()]
                
                transacoes_selecionadas = st.multiselect("Selecionar transações para excluir:", opcoes_transacao, key="bulk_delete_transacao_select")
                
                if transacoes_selecionadas:
                    st.warning(f"⚠️ Você está prestes a excluir {len(transacoes_selecionadas)} transação(ões). Esta ação não pode ser desfeita!")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("🗑️ Confirmar Exclusão em Lote", key="confirm_bulk_delete_transacao", type="primary"):
                            success_count = 0
                            error_count = 0
                            
                            for transacao_selecionada in transacoes_selecionadas:
                                idx_selecionado = opcoes_transacao.index(transacao_selecionada)
                                original_idx = df_bulk_delete.index[idx_selecionado]
                                success, message = crud_system.delete_record(sheet_name, original_idx)
                                
                                if success:
                                    success_count += 1
                                else:
                                    error_count += 1
                            
                            if success_count > 0:
                                st.success(f"✅ {success_count} transação(ões) excluída(s) com sucesso!")
                            if error_count > 0:
                                st.error(f"❌ {error_count} transação(ões) não puderam ser excluída(s).")
                            
                            st.session_state.show_bulk_delete_transacao = False
                            st.rerun()
                    
                    with col2:
                        if st.button("✖️ Cancelar", key="cancel_bulk_delete_transacao"):
                            st.session_state.show_bulk_delete_transacao = False
                            st.rerun()
            else:
                st.info(f"Nenhuma {tipo_transacao.lower()} encontrada com os filtros selecionados.")
        else:
            st.info("Nenhuma transação encontrada com os filtros selecionados.")

    elif selected == "Para onde vai":
        st.markdown('### Gráfico de Pareto das Despesas por Categoria')
        df_pareto = despesas_filtradas.copy()
        if not df_pareto.empty:
            pareto = (
                df_pareto.groupby('CATEGORIA')['VALOR']
                .sum()
                .abs()
                .sort_values(ascending=False)
                .reset_index()
            )
            pareto['%'] = pareto['VALOR'] / pareto['VALOR'].sum() * 100
            pareto['% Acumulado'] = pareto['%'].cumsum()
            pareto['Classe'] = 'C'
            pareto.loc[pareto['% Acumulado'] <= 80, 'Classe'] = 'A'
            pareto.loc[(pareto['% Acumulado'] > 80) & (pareto['% Acumulado'] <= 90), 'Classe'] = 'B'
            
            # Define as cores com base na classe ABC
            cores = pareto['Classe'].map({'A': '#fc6076', 'B': '#ffd200', 'C': '#b0b0b0'}).tolist()
            
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=pareto['CATEGORIA'],
                y=pareto['VALOR'],
                marker_color=cores, # Restaura as cores das barras
                name='Despesas',
                text=pareto['VALOR'].apply(lambda x: format_brl(x)),
                textposition='auto',
                width=0.4  # barras mais finas
            ))
            fig.add_trace(go.Scatter(
                x=pareto['CATEGORIA'],
                y=pareto['% Acumulado'],
                mode='lines+markers',
                name='Acumulado (%)',
                line=dict(color='#FFD700', width=3, dash='solid'), # Linha dourada
                marker=dict(color='#FFD700', size=8),
                yaxis='y2'
            ))
            fig.add_trace(go.Scatter(
                x=pareto['CATEGORIA'],
                y=[80]*len(pareto),
                mode='lines',
                name='80%',
                line=dict(color='#FFD700', width=2, dash='dash'), # Linha do limite de 80%
                yaxis='y2',
                showlegend=True
            ))
            fig.update_layout(
                yaxis=dict(title='Despesas (R$)', showgrid=False),
                yaxis2=dict(title='% Acumulado', overlaying='y', side='right', range=[0, 110], showgrid=False),
                xaxis=dict(title='Categoria', automargin=True, tickangle=-45, tickfont=dict(size=12)),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
                bargap=0.2,
                margin=dict(t=40, b=120),
                height=500,
                plot_bgcolor='rgba(0,0,0,0)', # Fundo transparente para se adaptar ao tema
                paper_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info('Não há despesas suficientes para exibir o gráfico de Pareto com os filtros selecionados.')

    elif selected == "Pra quem vai":
        st.markdown('### Gráfico de Pareto das Despesas por Descrição')
        df_pareto = despesas_filtradas.copy()
        if not df_pareto.empty:
            pareto = (
                df_pareto.groupby('DESCRIÇÃO')['VALOR']
                .sum()
                .abs()
                .sort_values(ascending=False)
                .reset_index()
            )
            pareto['%'] = pareto['VALOR'] / pareto['VALOR'].sum() * 100
            pareto['% Acumulado'] = pareto['%'].cumsum()
            pareto['Classe'] = 'C'
            pareto.loc[pareto['% Acumulado'] <= 80, 'Classe'] = 'A'
            pareto.loc[(pareto['% Acumulado'] > 80) & (pareto['% Acumulado'] <= 90), 'Classe'] = 'B'
            
            # Define as cores com base na classe ABC
            cores = pareto['Classe'].map({'A': '#fc6076', 'B': '#ffd200', 'C': '#b0b0b0'}).tolist()
            
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=pareto['DESCRIÇÃO'],
                y=pareto['VALOR'],
                marker_color=cores, # Restaura as cores das barras
                name='Despesas',
                text=pareto['VALOR'].apply(lambda x: format_brl(x)),
                textposition='auto',
                width=0.4
            ))
            fig.add_trace(go.Scatter(
                x=pareto['DESCRIÇÃO'],
                y=pareto['% Acumulado'],
                mode='lines+markers',
                name='Acumulado (%)',
                line=dict(color='#FFD700', width=3, dash='solid'), # Linha dourada
                marker=dict(color='#FFD700', size=8),
                yaxis='y2'
            ))
            fig.add_trace(go.Scatter(
                x=pareto['DESCRIÇÃO'],
                y=[80]*len(pareto),
                mode='lines',
                name='80%',
                line=dict(color='#FFD700', width=2, dash='dash'), # Linha do limite de 80%
                yaxis='y2',
                showlegend=True
            ))
            fig.update_layout(
                yaxis=dict(title='Despesas (R$)', showgrid=False),
                yaxis2=dict(title='% Acumulado', overlaying='y', side='right', range=[0, 110], showgrid=False),
                xaxis=dict(title='Descrição', automargin=True, tickangle=-45, tickfont=dict(size=12)),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
                bargap=0.2,
                margin=dict(t=40, b=120),
                height=500,
                plot_bgcolor='rgba(0,0,0,0)', # Fundo transparente para se adaptar ao tema
                paper_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info('Não há despesas suficientes para exibir o gráfico de Pareto com os filtros selecionados.')

    elif selected == "Classificação ABC":
        st.markdown('### Classificação ABC das Despesas por Descrição')
        df_abc = despesas_filtradas.copy()
        if not df_abc.empty:
            # Soma por descrição no mês atual
            desc_mes_atual = (
                df_abc.groupby('DESCRIÇÃO')['VALOR']
                .sum()
                .abs()
                .sort_values(ascending=False)
                .reset_index()
            )
            desc_mes_atual['%Despesas'] = desc_mes_atual['VALOR'] / desc_mes_atual['VALOR'].sum() * 100
            desc_mes_atual['Despesa Acumulada'] = desc_mes_atual['VALOR'].cumsum()
            desc_mes_atual['%Acum'] = desc_mes_atual['Despesa Acumulada'] / desc_mes_atual['VALOR'].sum() * 100
            desc_mes_atual['Ranking'] = desc_mes_atual['VALOR'].rank(ascending=False, method='min').astype(int)
            desc_mes_atual['Classe'] = 'C'
            desc_mes_atual.loc[desc_mes_atual['%Acum'] <= 80, 'Classe'] = 'A'
            desc_mes_atual.loc[(desc_mes_atual['%Acum'] > 80) & (desc_mes_atual['%Acum'] <= 90), 'Classe'] = 'B'
            # Soma por descrição no mês anterior
            if len(meses_selecionados) == 1:
                idx_mes = meses_ordem.index(meses_selecionados[0])
                if idx_mes > 0:
                    mes_ant = meses_ordem[idx_mes-1]
                    df_ant = despesas[(despesas['Ano'].isin(anos_selecionados)) & (despesas['Mês'] == mes_ant)]
                    desc_mes_ant = df_ant.groupby('DESCRIÇÃO')['VALOR'].sum().abs().reset_index()
                    desc_mes_ant = desc_mes_ant.rename(columns={'VALOR': 'Mês Anterior'})
                else:
                    desc_mes_ant = desc_mes_atual[['DESCRIÇÃO']].copy()
                    desc_mes_ant['Mês Anterior'] = 0
            else:
                desc_mes_ant = desc_mes_atual[['DESCRIÇÃO']].copy()
                desc_mes_ant['Mês Anterior'] = 0
            # Junta mês anterior
            df_abc = desc_mes_atual.merge(desc_mes_ant, on='DESCRIÇÃO', how='left')
            df_abc['Mês Anterior'] = df_abc['Mês Anterior'].fillna(0)
            # MoM e MoM%
            df_abc['MoM'] = df_abc['VALOR'] - df_abc['Mês Anterior']
            df_abc['MoM%_num'] = df_abc.apply(lambda x: (x['MoM']/x['Mês Anterior']*100) if x['Mês Anterior'] > 0 else 0, axis=1)
            # Formatação
            df_abc['Despesas(R$)'] = df_abc['VALOR'].apply(format_brl)
            df_abc['Mês Anterior'] = df_abc['Mês Anterior'].apply(format_brl)
            df_abc['MoM'] = df_abc['MoM'].apply(format_brl)
            df_abc['%Despesas'] = df_abc['%Despesas'].apply(lambda x: f'{x:.1f}%')
            df_abc['Despesa Acumulada'] = df_abc['Despesa Acumulada'].apply(format_brl)
            df_abc['%Acum'] = df_abc['%Acum'].apply(lambda x: f'{x:.1f}%')
            # MoM% com setas
            mom_seta = []
            for val in df_abc['MoM%_num']:
                if val > 0:
                    seta = '<span class="seta-verde">▲</span>'
                elif val < 0:
                    seta = '<span class="seta-vermelha">▼</span>'
                else:
                    seta = '—'
                mom_seta.append(f'{val:.1f}% {seta}')
            df_abc['MoM%'] = mom_seta
            # Classe com ícones
            classe_icone = []
            for classe in df_abc['Classe']:
                if classe == 'A':
                    icone = '<span class="seta-vermelha">▼</span>'
                elif classe == 'B':
                    icone = '<span class="classe-b-icone">●</span>'
                else:
                    icone = '<span class="seta-cinza">▼</span>'
                classe_icone.append(f'{classe} {icone}')
            df_abc['Classe_Icone'] = classe_icone
            # Seleção de colunas e renomeação
            df_abc = df_abc.rename(columns={'DESCRIÇÃO': 'Descrição'})
            df_abc = df_abc[['Descrição', 'Ranking', 'Despesas(R$)', '%Despesas', 'Despesa Acumulada', '%Acum', 'Classe_Icone', 'Mês Anterior', 'MoM', 'MoM%']]
            # Renderização
            st.markdown('#### Tabela ABC detalhada das despesas')
            # Tabela HTML customizada para colorir Descrição
            table_html = """
            <style>
            .abc-table {
                width: 100%;
                border-collapse: collapse;
                color: #f0f2f6; /* Texto claro para tema escuro */
                font-family: 'sans-serif';
            }
            .abc-table th, .abc-table td {
                border: 1px solid #3a3f44;
                padding: 8px 12px;
                text-align: center;
            }
            .abc-table th {
                background-color: #1a1f24; /* Cabeçalho mais escuro */
            }
            .abc-table tr:nth-child(even) {
                background-color: #262c31; /* Listras para melhor leitura */
            }
            .cat-a { color: #ff6961; font-weight: bold; } /* Vermelho para Classe A */
            .cat-b { color: #fdfd96; font-weight: bold; } /* Amarelo para Classe B */
            .seta-verde { color: #77dd77; font-weight: bold; } /* Verde suave */
            .seta-vermelha { color: #ff6961; font-weight: bold; } /* Vermelho suave */
            .seta-cinza { color: #8390a2; font-weight: bold; } /* Cinza azulado */
            .classe-b-icone { color: #fdfd96; font-weight: bold; } /* Amarelo para Classe B */
            </style>
            <table class='abc-table'>
                <tr>
                    <th>Descrição</th>
                    <th>Ranking</th>
                    <th>Despesas(R$)</th>
                    <th>%Despesas</th>
                    <th>Despesa Acumulada</th>
                    <th>%Acum</th>
                    <th>Classe</th>
                    <th>Mês Anterior</th>
                    <th>MoM</th>
                    <th>MoM%</th>
                </tr>
            """
            for _, row in df_abc.iterrows():
                classe = row['Classe_Icone']
                if 'A' in classe:
                    desc_class = 'cat-a'
                elif 'B' in classe:
                    desc_class = 'cat-b'
                else:
                    desc_class = ''
                table_html += f"""
                <tr>
                    <td class='{desc_class}'>{row['Descrição']}</td>
                    <td>{row['Ranking']}</td>
                    <td>{row['Despesas(R$)']}</td>
                    <td>{row['%Despesas']}</td>
                    <td>{row['Despesa Acumulada']}</td>
                    <td>{row['%Acum']}</td>
                    <td>{row['Classe_Icone']}</td>
                    <td>{row['Mês Anterior']}</td>
                    <td>{row['MoM']}</td>
                    <td>{row['MoM%']}</td>
                </tr>
                """
            table_html += "</table>"
            components.html(table_html, height=600, scrolling=True)
        else:
            st.info('Não há despesas suficientes para exibir a tabela ABC com os filtros selecionados.')

    elif selected == "Fluxo de Caixa":
        st.markdown('### Fluxo de Caixa')
        # Junta receitas e despesas filtradas
        df_fluxo = pd.concat([
            receitas_filtradas.assign(TIPO='Receita'),
            despesas_filtradas.assign(TIPO='Despesa')
        ], ignore_index=True)
        df_fluxo = df_fluxo.sort_values('DATA')
        df_fluxo['VALOR'] = df_fluxo.apply(lambda x: x['VALOR'] if x['TIPO']=='Receita' else x['VALOR'], axis=1)
        df_fluxo['Saldo Acumulado'] = df_fluxo['VALOR'].cumsum()
        df_fluxo['DATA'] = df_fluxo['DATA'].dt.strftime('%d/%m/%Y')
        # Gráfico de área customizado estilo dark
        if not df_fluxo.empty:
            fig_fluxo = go.Figure()
            fig_fluxo.add_trace(go.Scatter(
                x=df_fluxo['DATA'],
                y=df_fluxo['Saldo Acumulado'],
                mode='lines+markers+text',
                text=[f"R$ {v:,.0f}".replace(",", ".") for v in df_fluxo['Saldo Acumulado']],
                textposition='top center',
                hovertemplate='Data: %{x}<br>Saldo: R$ %{y:,.2f}<extra></extra>',
                name='Saldo Acumulado'
            ))
            fig_fluxo.update_layout(
                title='Análise do Saldo Acumulado (YTD)',
                xaxis=dict(title='Data'),
                yaxis=dict(title='Saldo Acumulado (R$)'),
                margin=dict(l=30, r=30, t=50, b=40),
                height=420,
                showlegend=False,
            )
            st.plotly_chart(fig_fluxo, use_container_width=True)
        else:
            st.info('Não há dados suficientes para exibir o fluxo de caixa com os filtros selecionados.')

    elif selected == "Despesas por Categoria":
        st.markdown('### Despesas por Subcategoria dentro da Categoria')
        df_cat = despesas_filtradas.copy()
        if not df_cat.empty:
            if categoria_selecionada == 'Todas':
                st.info('Selecione uma categoria específica na barra lateral para ver o detalhamento por subcategoria.')
            else:
                df_cat = df_cat[df_cat['CATEGORIA'] == categoria_selecionada]
                subcats = df_cat['DESCRIÇÃO'].dropna().unique().tolist()
                subcats = sorted(subcats)
                subcats_opcoes = ['Todas'] + subcats
                subcats_selecionadas = st.multiselect('Filtrar Subcategoria', subcats_opcoes, default=['Todas'])
                if 'Todas' in subcats_selecionadas or not subcats_selecionadas:
                    df_sub = df_cat
                    subcats_plot = subcats
                else:
                    df_sub = df_cat[df_cat['DESCRIÇÃO'].isin(subcats_selecionadas)]
                    subcats_plot = subcats_selecionadas
                # Agrupa por mês e subcategoria
                df_sub['MÊS_ANO'] = df_sub['DATA'].dt.strftime('%b/%Y').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
                evolucao = df_sub.groupby(['MÊS_ANO', 'DESCRIÇÃO'])['VALOR'].sum().abs().reset_index()
                # Ordena os meses corretamente
                meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                evolucao['MÊS'] = evolucao['MÊS_ANO'].str[:3]
                evolucao['ANO'] = evolucao['MÊS_ANO'].str[-4:]
                evolucao['MÊS_NUM'] = evolucao['MÊS'].apply(lambda x: meses_ordem.index(x) if x in meses_ordem else -1)
                evolucao['ANO_NUM'] = evolucao['ANO'].astype(int)
                evolucao = evolucao.sort_values(['ANO_NUM', 'MÊS_NUM'])
                # Gera lista completa de meses/anos entre o menor e o maior registro
                if not evolucao.empty:
                    min_ano = evolucao['ANO_NUM'].min()
                    max_ano = evolucao['ANO_NUM'].max()
                    min_mes = evolucao[evolucao['ANO_NUM'] == min_ano]['MÊS_NUM'].min()
                    max_mes = evolucao[evolucao['ANO_NUM'] == max_ano]['MÊS_NUM'].max()
                    meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                    labels_x = []
                    for ano in range(min_ano, max_ano+1):
                        for i, mes in enumerate(meses_ordem):
                            if (ano == min_ano and i < min_mes) or (ano == max_ano and i > max_mes):
                                continue
                            labels_x.append(f'{mes}/{ano}')
                else:
                    labels_x = []
                # Gráfico de linha
                fig = go.Figure()
                for subcat in subcats_plot:
                    dados = evolucao[evolucao['DESCRIÇÃO'] == subcat]
                    # Garante que todos os meses apareçam, mesmo se não houver valor
                    dados = dados.set_index('MÊS_ANO').reindex(labels_x, fill_value=0).reset_index()
                    fig.add_trace(go.Scatter(
                        x=dados['MÊS_ANO'],
                        y=dados['VALOR'],
                        mode='lines+markers',
                        name=subcat,
                        text=[f"R$ {v:,.2f}".replace(",", ".") for v in dados['VALOR']],
                        textposition='top center',
                    ))
                # Linha de média
                if not evolucao.empty:
                    media = evolucao.groupby('MÊS_ANO')['VALOR'].sum().mean()
                    fig.add_trace(go.Scatter(
                        x=labels_x,
                        y=[media]*len(labels_x),
                        mode='lines',
                        name='Média',
                    ))
                fig.update_layout(
                    title=f'Evolução Mensal das Subcategorias em {categoria_selecionada}',
                    xaxis_title='Mês/Ano',
                    yaxis_title='Valor Gasto (R$)',
                    height=480,
                    margin=dict(l=30, r=30, t=50, b=120),
                    xaxis=dict(
                        tickangle=-35,
                        tickfont=dict(size=13),
                        categoryorder='array',
                        categoryarray=labels_x,
                        type='category'
                    ),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
                )
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.info('Não há despesas para exibir por subcategoria com os filtros selecionados.')

    # --- GRUPO DE ANÁLISE DE VENDAS ---
    if selected == "Vendas":
        st.markdown('## 📈 Análise de Vendas')
        
        try:
            df_vendas = pd.read_excel(xls, sheet_name='Vendas')
            df_vendas['DATA'] = pd.to_datetime(df_vendas['DATA'], errors='coerce')
            df_vendas.dropna(subset=['DATA'], inplace=True)
            
            # DEBUG Adicional para ver o que foi removido
            if len(df_vendas) == 0 and 'DATA' in df_vendas.columns:
                 st.warning("Nenhuma venda encontrada com data válida. Verifique a coluna 'DATA' na sua planilha 'Vendas'.")

            df_vendas['Ano'] = df_vendas['DATA'].dt.year.astype(str)
            df_vendas['Mês'] = df_vendas['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            vendas_filtradas = df_vendas[
                (df_vendas['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
                (df_vendas['Mês'].isin(meses_selecionados) if meses_selecionados else True) &
                ((df_vendas['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas'))
            ]

        except Exception as e:
            st.error(f"Ocorreu um erro ao carregar os dados de Vendas: {e}")
            vendas_filtradas = pd.DataFrame() # Cria um dataframe vazio para evitar mais erros

        # Indicadores
        total_vendido = vendas_filtradas['VALOR'].sum() if not vendas_filtradas.empty else 0
        num_vendas = len(vendas_filtradas)
        ticket_medio = total_vendido / num_vendas if num_vendas > 0 else 0
        total_recebido = vendas_filtradas[vendas_filtradas['Status'].astype(str).str.lower() == 'sim']['VALOR'].sum() if not vendas_filtradas.empty else 0
        total_a_receber = vendas_filtradas[vendas_filtradas['Status'].astype(str).str.lower() == 'não']['VALOR'].sum() if not vendas_filtradas.empty else 0

        st.markdown('---')
        # Indicadores de vendas com cartões escuros padronizados
        colv1, colv2, colv3, colv4, colv5 = st.columns(5)
        with colv1:
            st.metric("Total Vendido", format_brl(total_vendido))
        with colv2:
            st.metric("Nº de Vendas", num_vendas)
        with colv3:
            st.metric("Ticket Médio", format_brl(ticket_medio))
        with colv4:
            st.metric("Total Recebido", format_brl(total_recebido))
        with colv5:
            st.metric("Total a Receber", format_brl(total_a_receber))

        # Gráficos de rosca lado a lado
        st.markdown('---')
        colg1, colg2 = st.columns(2)
        with colg1:
            # Vendas por cliente (top 5 em valor)
            if not vendas_filtradas.empty:
                top_clientes = vendas_filtradas.groupby('Cliente')['VALOR'].sum().sort_values(ascending=False).head(5)
                fig_cli = px.pie(names=top_clientes.index, values=top_clientes.values, hole=0.5, title='Top 5 Clientes (Valor)')
                fig_cli.update_traces(textinfo='percent', textposition='outside', pull=[0.05]*5, textfont=dict(family='Arial', size=16, color='white'), textfont_weight='bold')
                fig_cli.update_layout(
                    title_font_size=22,
                    legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5, font=dict(size=13))
                )
                st.plotly_chart(fig_cli, use_container_width=True)
                # Top 5 clientes em quantidade de vendas
                top_clientes_qtd = vendas_filtradas['Cliente'].value_counts().head(5)
                fig_cli_qtd = px.pie(names=top_clientes_qtd.index, values=top_clientes_qtd.values, hole=0.5, title='Top 5 Clientes (Qtd. Vendas)')
                fig_cli_qtd.update_traces(textinfo='percent', textposition='outside', pull=[0.05]*5, textfont=dict(family='Arial', size=16, color='white'), textfont_weight='bold')
                fig_cli_qtd.update_layout(
                    title_font_size=22,
                    legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5, font=dict(size=13))
                )
                st.plotly_chart(fig_cli_qtd, use_container_width=True)
            else:
                st.info('Não há vendas para exibir por cliente.')
        with colg2:
            # Vendas por tipo de recebimento
            if not vendas_filtradas.empty:
                tipo_receb = vendas_filtradas.groupby('TIPO DE RECEBIMENTO')['VALOR'].sum().sort_values(ascending=False)
                fig_tipo = px.pie(names=tipo_receb.index, values=tipo_receb.values, hole=0.5, title='Por Tipo de Recebimento')
                fig_tipo.update_traces(textinfo='percent', textposition='outside', textfont=dict(family='Arial', size=16, color='white'), textfont_weight='bold')
                fig_tipo.update_layout(
                    title_font_size=22,
                    legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5, font=dict(size=13))
                )
                st.plotly_chart(fig_tipo, use_container_width=True)
            else:
                st.info('Não há vendas para exibir por tipo de recebimento.')

        # Gráfico de evolução das vendas (abaixo dos roscas)
        st.markdown('---')
        st.markdown('#### Evolução Mensal das Vendas')
        if not vendas_filtradas.empty:
            evol = vendas_filtradas.groupby(['Ano', 'Mês'])['VALOR'].sum().reset_index()
            evol['MÊS_NUM'] = evol['Mês'].apply(lambda x: meses_ordem.index(x) if x in meses_ordem else -1)
            evol = evol.sort_values(['Ano', 'MÊS_NUM'])
            evol['MêsAno'] = evol['Mês'] + '/' + evol['Ano']
            fig_evol = px.bar(
                evol,
                x='MêsAno',
                y='VALOR',
                labels={'MêsAno':'Mês/Ano','VALOR':'Total Vendido'},
                text=evol['VALOR'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "."))
            )
            fig_evol.update_traces(textposition='auto', textfont_size=14)
            fig_evol.update_layout(xaxis_tickangle=-35, height=320, margin=dict(t=60, b=40, l=0, r=0))
            st.plotly_chart(fig_evol, use_container_width=True)
        else:
            st.info('Não há vendas para exibir evolução.')

        # Tabela detalhada
        st.markdown('---')
        st.markdown('### Tabela Detalhada de Vendas')
        if not vendas_filtradas.empty:
            df_tab = vendas_filtradas.copy()
            df_tab['DATA'] = df_tab['DATA'].dt.strftime('%d/%m/%Y')
            df_tab = df_tab.rename(columns={'DATA':'Data','CONTA':'Conta','TIPO DE RECEBIMENTO':'Tipo de Recebimento','VALOR':'Valor','Status':'Status'})
            df_tab['Valor'] = df_tab['Valor'].apply(format_brl)
            st.dataframe(df_tab[['Data','Cliente','Conta','Tipo de Recebimento','Valor','Status']], use_container_width=True, hide_index=True)
            
            # Botões de ação CRUD para Vendas
            st.markdown('---')
            col_crud1, col_crud2, col_crud3, col_crud4, col_crud5, col_crud6 = st.columns([1, 1, 1, 1, 1, 1])
            
            with col_crud1:
                if st.button("🛒 Nova Venda", key="nova_venda"):
                    st.session_state.show_venda_form = True
            
            with col_crud2:
                if st.button("✏️ Editar Venda", key="edit_venda"):
                    st.session_state.show_edit_Vendas = True
            
            with col_crud3:
                if st.button("🗑️ Excluir Venda", key="delete_venda"):
                    st.session_state.show_delete_Vendas = True
            
            with col_crud4:
                if st.button("🗑️ Exclusão em Lote", key="bulk_delete_venda"):
                    st.session_state.show_bulk_delete_Vendas = True
        else:
            st.info('Não há vendas para exibir na tabela.')

    # --- GRUPO DE ANÁLISE DE INVESTIMENTOS ---
    if selected == "Investimentos":
        st.markdown('## 💰 Análise de Investimentos')
        
        # Carrega dados de investimentos e metas
        df_investimentos = pd.read_excel(xls, sheet_name='Investimentos')
        df_metas = pd.read_excel(xls, sheet_name='Metas')
        
        # Processa dados de investimentos
        df_investimentos['DATA'] = pd.to_datetime(df_investimentos['DATA'], errors='coerce')
        df_investimentos = df_investimentos.dropna(subset=['DATA'])
        df_investimentos['Ano'] = df_investimentos['DATA'].dt.year.astype(str)
        df_investimentos['Mês'] = df_investimentos['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
        
        # Aplica filtros
        investimentos_filtrados = df_investimentos[
            (df_investimentos['Ano'].isin(st.session_state.get('ano_invest', [])) if st.session_state.get('ano_invest') else True) &
            (df_investimentos['Mês'].isin(st.session_state.get('mes_invest', [])) if st.session_state.get('mes_invest') else True) &
            ((df_investimentos['TIPO'] == st.session_state.get('tipo_invest', 'Todos')) | (st.session_state.get('tipo_invest', 'Todos') == 'Todos')) &
            ((df_investimentos['OBJETIVO'] == st.session_state.get('objetivo_invest', 'Todos')) | (st.session_state.get('objetivo_invest', 'Todos') == 'Todos')) &
            ((df_investimentos['ATIVO'] == st.session_state.get('ativo_invest', 'Todos')) | (st.session_state.get('ativo_invest', 'Todos') == 'Todos'))
        ]
        
        # Calcula indicadores
        total_investido = investimentos_filtrados['VALOR_APORTE'].sum()
        num_aportes = len(investimentos_filtrados)
        valor_medio_aporte = total_investido / num_aportes if num_aportes > 0 else 0
        
        # Calcula progresso das metas
        ano_atual = datetime.now().year
        meta_atual = df_metas[df_metas['ANO'] == ano_atual]['META_TOTAL'].sum()
        progresso_meta = (total_investido / meta_atual * 100) if meta_atual > 0 else 0
        falta_para_meta = meta_atual - total_investido if meta_atual > total_investido else 0
        
        # Exibe indicadores principais
        st.markdown('---')
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("Total Investido", format_brl(total_investido))
        with col2:
            st.metric("Nº de Aportes", num_aportes)
        with col3:
            st.metric("Valor Médio", format_brl(valor_medio_aporte))
        with col4:
            st.metric("Meta Anual", format_brl(meta_atual))
        with col5:
            st.metric("Progresso", f"{progresso_meta:.1f}%")
        
        # Barra de progresso da meta
        st.markdown('---')
        st.markdown('### 📊 Progresso da Meta Anual')
        col_prog1, col_prog2 = st.columns([3, 1])
        with col_prog1:
            st.progress(progresso_meta / 100)
        with col_prog2:
            if falta_para_meta > 0:
                st.metric("Falta", format_brl(falta_para_meta), delta="Para meta")
            else:
                st.metric("Meta Atingida! 🎉", format_brl(0), delta="Parabéns!")
        
        # Gráficos de análise
        st.markdown('---')
        col_graf1, col_graf2 = st.columns(2)
        
        with col_graf1:
            # Distribuição por tipo de investimento
            if not investimentos_filtrados.empty:
                tipo_dist = investimentos_filtrados.groupby('TIPO')['VALOR_APORTE'].sum().sort_values(ascending=False)
                fig_tipo = px.pie(
                    names=tipo_dist.index, 
                    values=tipo_dist.values, 
                    hole=0.5, 
                    title='Distribuição por Tipo de Investimento',
                )
                fig_tipo.update_traces(
                    textinfo='percent+label', 
                    textposition='outside',
                    textfont=dict(family='Arial', size=14, color='white')
                )
                fig_tipo.update_layout(
                    title_font_size=20,
                    legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5)
                )
                st.plotly_chart(fig_tipo, use_container_width=True)
            else:
                st.info('Não há investimentos para exibir.')
        
        with col_graf2:
            # Distribuição por objetivo
            if not investimentos_filtrados.empty:
                obj_dist = investimentos_filtrados.groupby('OBJETIVO')['VALOR_APORTE'].sum().sort_values(ascending=False)
                fig_obj = px.pie(
                    names=obj_dist.index, 
                    values=obj_dist.values, 
                    hole=0.5, 
                    title='Distribuição por Objetivo',
                )
                fig_obj.update_traces(
                    textinfo='percent+label', 
                    textposition='outside',
                    textfont=dict(family='Arial', size=14, color='white')
                )
                fig_obj.update_layout(
                    title_font_size=20,
                    legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5)
                )
                st.plotly_chart(fig_obj, use_container_width=True)
            else:
                st.info('Não há investimentos para exibir.')
        
        # Gráfico de evolução mensal
        st.markdown('---')
        st.markdown('### 📈 Evolução Mensal dos Aportes')
        if not investimentos_filtrados.empty:
            evol_mensal = investimentos_filtrados.groupby(['Ano', 'Mês'])['VALOR_APORTE'].sum().reset_index()
            evol_mensal['MÊS_NUM'] = evol_mensal['Mês'].apply(lambda x: meses_ordem.index(x) if x in meses_ordem else -1)
            evol_mensal = evol_mensal.sort_values(['Ano', 'MÊS_NUM'])
            evol_mensal['MêsAno'] = evol_mensal['Mês'] + '/' + evol_mensal['Ano']
            
            fig_evol = px.bar(
                evol_mensal,
                x='MêsAno',
                y='VALOR_APORTE',
                labels={'MêsAno':'Mês/Ano','VALOR_APORTE':'Total Aportado'},
                text=evol_mensal['VALOR_APORTE'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "."))
            )
            fig_evol.update_traces(textposition='auto', textfont_size=12)
            fig_evol.update_layout(
                xaxis_tickangle=-35, 
                height=400, 
                margin=dict(t=60, b=40, l=0, r=0),
                title='Evolução dos Aportes por Mês'
            )
            st.plotly_chart(fig_evol, use_container_width=True)
        else:
            st.info('Não há dados suficientes para exibir a evolução mensal.')
        
        # Top 10 ativos mais investidos
        st.markdown('---')
        st.markdown('### 🏆 Top 10 Ativos Mais Investidos')
        if not investimentos_filtrados.empty:
            top_ativos = investimentos_filtrados.groupby('ATIVO')['VALOR_APORTE'].sum().sort_values(ascending=False).head(10)
            fig_top = px.bar(
                x=top_ativos.values,
                y=top_ativos.index,
                orientation='h',
                labels={'x':'Valor Investido (R$)', 'y':'Ativo'},
                text=[f"R$ {x:,.2f}".replace(",", ".") for x in top_ativos.values]
            )
            fig_top.update_traces(textposition='auto', textfont_size=11)
            fig_top.update_layout(
                height=500,
                margin=dict(t=40, b=40, l=0, r=0),
                title='Top 10 Ativos por Valor Investido'
            )
            st.plotly_chart(fig_top, use_container_width=True)
        else:
            st.info('Não há dados suficientes para exibir o ranking de ativos.')
        
        # Tabela detalhada de investimentos
        st.markdown('---')
        st.markdown('### 📋 Tabela Detalhada de Investimentos')
        if not investimentos_filtrados.empty:
            df_tab = investimentos_filtrados.copy()
            df_tab['DATA'] = df_tab['DATA'].dt.strftime('%d/%m/%Y')
            df_tab = df_tab.rename(columns={
                'DATA':'Data',
                'TIPO':'Tipo',
                'ATIVO':'Ativo',
                'VALOR_APORTE':'Valor Aportado',
                'QUANTIDADE':'Quantidade',
                'PRECO_MEDIO':'Preço Médio',
                'OBJETIVO':'Objetivo'
            })
            df_tab['Valor Aportado'] = df_tab['Valor Aportado'].apply(format_brl)
            df_tab['Preço Médio'] = df_tab['Preço Médio'].apply(format_brl)
            st.dataframe(
                df_tab[['Data','Tipo','Ativo','Valor Aportado','Quantidade','Preço Médio','Objetivo']], 
                use_container_width=True,
                hide_index=True
            )
            
            # Botões de ação CRUD para Investimentos
            st.markdown('---')
            col_crud1, col_crud2, col_crud3, col_crud4 = st.columns([1, 1, 1, 1])
            
            with col_crud1:
                if st.button("💰 Novo Investimento", key="novo_investimento"):
                    st.session_state.show_investimento_form = True
            
            with col_crud2:
                if st.button("✏️ Editar Investimento", key="edit_investimento"):
                    st.session_state.show_edit_Investimentos = True
            
            with col_crud3:
                if st.button("🗑️ Excluir Investimento", key="delete_investimento"):
                    st.session_state.show_delete_Investimentos = True
            
            with col_crud4:
                if st.button("🗑️ Exclusão em Lote", key="bulk_delete_investimento"):
                    st.session_state.show_bulk_delete_Investimentos = True
        else:
            st.info('Não há investimentos para exibir na tabela.')

    # --- GRUPO DE ANÁLISE DE CARTÃO DE CRÉDITO ---
    elif selected == "Cartão de Crédito":
        st.markdown('## 💳 Análise de Cartão de Crédito')

        # Carregar e processar dados de CC
        try:
            df_cc = pd.read_excel(xls, sheet_name='Div_CC')
            if not df_cc.empty:
                df_cc['Data'] = pd.to_datetime(df_cc['Data'], errors='coerce')
                df_cc.dropna(subset=['Data'], inplace=True)
                df_cc['Ano'] = df_cc['Data'].dt.year.astype(str)
                df_cc['Mês'] = df_cc['Data'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})

                # Exibe os filtros na sidebar
                st.sidebar.header("Filtros Cartão de Crédito")
                
                anos_disponiveis = sorted(df_cc['Ano'].unique())
                anos_selecionados_cc = st.sidebar.multiselect('Ano', anos_disponiveis, default=anos_disponiveis, key='cc_ano_multiselect')
                
                meses_disponiveis = sorted(df_cc['Mês'].unique())
                meses_selecionados_cc = st.sidebar.multiselect('Mês', meses_disponiveis, default=[], key='cc_mes_multiselect')
                
                cartoes_disponiveis = ['Todos'] + sorted(df_cc['Cartão'].unique().tolist())
                cartao_selecionado = st.sidebar.selectbox('Cartão', cartoes_disponiveis, key='cc_cartao_select')
                
                situacoes_disponiveis = ['Todas'] + sorted(df_cc['Situação'].unique().tolist())
                situacao_selecionada = st.sidebar.selectbox('Situação', situacoes_disponiveis, key='cc_situacao_select')
                
                tipos_compra_disponiveis = ['Todos'] + sorted(df_cc['Tipo de Compra'].unique().tolist())
                tipo_compra_selecionado = st.sidebar.selectbox('Tipo de Compra', tipos_compra_disponiveis, key='cc_tipo_compra_select')

                # Aplicar filtros
                df_cc_filtrado = df_cc.copy()
                if anos_selecionados_cc:
                    df_cc_filtrado = df_cc_filtrado[df_cc_filtrado['Ano'].isin(anos_selecionados_cc)]
                if meses_selecionados_cc:
                    df_cc_filtrado = df_cc_filtrado[df_cc_filtrado['Mês'].isin(meses_selecionados_cc)]
                if cartao_selecionado != 'Todos':
                    df_cc_filtrado = df_cc_filtrado[df_cc_filtrado['Cartão'] == cartao_selecionado]
                if situacao_selecionada != 'Todas':
                    df_cc_filtrado = df_cc_filtrado[df_cc_filtrado['Situação'] == situacao_selecionada]
                if tipo_compra_selecionado != 'Todos':
                    df_cc_filtrado = df_cc_filtrado[df_cc_filtrado['Tipo de Compra'] == tipo_compra_selecionado]

                # Cards de resumo
                total_gasto_cc = df_cc_filtrado['valor total da compra'].sum()
                proxima_fatura = df_cc_filtrado[df_cc_filtrado['Situação'] == 'Pendente']['valor total da compra'].sum()
                media_mensal = total_gasto_cc / (len(anos_selecionados_cc) * 12) if anos_selecionados_cc else 0

                col1, col2, col3 = st.columns(3)
                col1.metric("Total Gasto (Filtrado)", format_brl(total_gasto_cc))
                col2.metric("Próxima Fatura (Pendente)", format_brl(proxima_fatura))
                col3.metric("Média Mensal (Filtrado)", format_brl(media_mensal))

                st.markdown("---")

                # Tabela de detalhes
                with st.container():
                    st.markdown("#### Detalhes das Compras")
                    
                    df_exibicao_cc = df_cc_filtrado[['Data', 'Descrição', 'valor total da compra', 'Situação', 'Cartão', 'Tipo de Compra', 'Quantidade de parcelas', 'id']].rename(columns={'valor total da compra': 'Valor', 'Quantidade de parcelas': 'Parcelas'})
                    
                    if not df_exibicao_cc.empty:
                        # Aplica a formatação de moeda na coluna 'Valor'
                        df_exibicao_cc['Valor'] = df_exibicao_cc['Valor'].apply(format_brl)
                        st.dataframe(df_exibicao_cc, use_container_width=True, hide_index=True)
                    else:
                        st.info('Não há compras para exibir na tabela com os filtros selecionados.')

                # Botões de ação CRUD e Relatório
                st.markdown('---')
                col_crud1, col_crud2, col_crud3, col_crud4, col_crud5 = st.columns([1.2, 1, 1, 1.2, 1.5])
                with col_crud1:
                    if st.button("💳 Nova Compra (CC)", key="nova_compra_cc_main", use_container_width=True):
                        st.session_state.show_cc_form = True
                with col_crud2:
                    if st.button("✏️ Editar", key="edit_compra_cc_main", use_container_width=True, disabled=df_exibicao_cc.empty):
                        st.session_state.show_edit_Div_CC = True
                with col_crud3:
                    if st.button("🗑️ Excluir", key="delete_compra_cc_main", use_container_width=True, disabled=df_exibicao_cc.empty):
                        st.session_state.show_delete_Div_CC = True
                with col_crud4:
                    if st.button("🗑️ Excluir em Lote", key="lote_compra_cc_main", use_container_width=True, disabled=df_exibicao_cc.empty):
                        st.session_state.show_delete_lote_cc = True
                with col_crud5:
                    if not df_exibicao_cc.empty:
                        try:
                            # Instancia o relatório e gera o PDF
                            relatorio = RelatorioPDF()
                            pdf_buffer = relatorio.gerar_pdf_compras(df_exibicao_cc, total_gasto_cc, proxima_fatura, media_mensal)
                            st.download_button(label="📄 Exportar Relatório PDF", data=pdf_buffer, file_name="relatorio_compras_cc.pdf", mime="application/pdf", use_container_width=True)
                        except Exception as e:
                            st.error(f"❌ Erro ao gerar PDF: {e}")
                    else:
                        st.button("📄 Exportar Relatório PDF", use_container_width=True, disabled=True)
            else:
                st.info("Não há dados de compras de cartão de crédito para exibir.")
                if st.button("💳 Nova Compra (CC)", key="nova_compra_cc_main_no_data"):
                    st.session_state.show_cc_form = True

        except FileNotFoundError:
            st.error("A aba 'Div_CC' não foi encontrada na planilha. Verifique o arquivo 'Base_financas.xlsx'.")
        except Exception as e:
            st.error(f"Ocorreu um erro ao carregar os dados do cartão de crédito: {e}")

    # --- GRUPO DE ANÁLISE DE ORÇAMENTO ---
    elif selected == "Orçamento":
        st.markdown("## 📊 Análise de Orçamento Mensal")

        try:
            df_orcamento = pd.read_excel(xls, sheet_name='Orcamento')
            if 'Categoria' not in df_orcamento.columns or 'Percentual' not in df_orcamento.columns:
                st.error("A aba 'Orcamento' deve conter as colunas 'Categoria' and 'Percentual'.")
                st.stop()
        except ValueError:
            st.warning("A aba 'Orcamento' não foi encontrada na sua planilha 'Base_financas.xlsx'.")
            st.info("Para usar esta funcionalidade, por favor, crie uma aba chamada 'Orcamento' com as colunas 'Categoria' e 'Percentual' (ex: Moradia, 30).")
            st.stop()

        # Garante que os filtros de período (mês/ano) sejam selecionados
        if not anos_selecionados or not meses_selecionados:
            st.info("👈 Por favor, selecione pelo menos um Ano e um Mês na barra lateral para calcular o orçamento.")
            st.stop()
            
        # 1. Filtra os dados de receita e despesa para o período selecionado, sem filtro de categoria
        receitas_periodo = receitas[
            ((receitas['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas')) &
            (receitas['Ano'].isin(anos_selecionados)) &
            (receitas['Mês'].isin(meses_selecionados))
        ]
        despesas_periodo = despesas[
            ((despesas['CONTA'] == conta_selecionada) | (conta_selecionada == 'Todas')) &
            (despesas['Ano'].isin(anos_selecionados)) &
            (despesas['Mês'].isin(meses_selecionados))
        ]

        # 2. Calcula a Renda Líquida Base para o orçamento
        total_receitas_periodo = receitas_periodo['VALOR'].sum()
        # Assume que 'Confeitaria' é um custo de negócio, a ser deduzido das receitas
        custos_negocio = despesas_periodo[despesas_periodo['CATEGORIA'] == 'Confeitaria']['VALOR'].sum()
        renda_liquida_base = total_receitas_periodo + custos_negocio # Custos já são negativos

        st.metric(
            "Renda Líquida para Orçamento (Receitas - Custos de 'Confeitaria')",
            format_brl(renda_liquida_base)
        )
        st.caption(f"Cálculo: {format_brl(total_receitas_periodo)} (Receitas) - {format_brl(abs(custos_negocio))} (Custos Confeitaria)")

        st.markdown("---")

        if renda_liquida_base > 0:
            # 3. Calcula os gastos reais por categoria (excluindo os custos de negócio)
            gastos_reais_cat = despesas_periodo[despesas_periodo['CATEGORIA'] != 'Confeitaria'].groupby('CATEGORIA')['VALOR'].sum().abs().reset_index()
            gastos_reais_cat = gastos_reais_cat.rename(columns={'VALOR': 'Gasto', 'CATEGORIA': 'Categoria'})

            # 4. Cria a tabela de análise do orçamento
            df_analise = df_orcamento.merge(gastos_reais_cat, on='Categoria', how='left')
            df_analise['Gasto'] = df_analise['Gasto'].fillna(0)
            df_analise['Orcado'] = (df_analise['Percentual'] / 100) * renda_liquida_base
            df_analise['Saldo'] = df_analise['Orcado'] - df_analise['Gasto']
            
            # Evita divisão por zero se o orçado for 0
            df_analise['%_Uso'] = (df_analise['Gasto'] / df_analise['Orcado'].replace(0, np.nan)) * 100
            df_analise['%_Uso'] = df_analise['%_Uso'].fillna(0)

            # 5. Exibe o progresso por categoria
            st.markdown("### Progresso por Categoria")
            for _, row in df_analise.sort_values(by='Percentual', ascending=False).iterrows():
                st.markdown(f"**{row['Categoria']}** ({row['Percentual']}% do orçamento)")
                col1, col2 = st.columns([3, 1])
                
                # Barra de progresso visual
                percent_usage_visual = min(row['%_Uso'] / 100, 1.0)
                prog_bar = col1.progress(percent_usage_visual)
                
                # Define a cor da barra de progresso
                if row['%_Uso'] > 100:
                    prog_bar.empty()
                    col1.error(f"Orçamento estourado em {format_brl(abs(row['Saldo']))}!")
                
                # Métrica de status
                delta_color = "normal" if row['Saldo'] >= 0 else "inverse"
                col2.metric(
                    label=f"Gasto de {format_brl(row['Gasto'])}",
                    value=f"{row['%_Uso']:.1f}%",
                    delta=f"Saldo: {format_brl(row['Saldo'])}",
                    delta_color=delta_color
                )
            
            st.markdown("---")

            # 6. Gráfico de Comparação
            st.markdown("### Gráfico: Orçado vs. Gasto")
            df_plot = df_analise.melt(
                id_vars=['Categoria'],
                value_vars=['Orcado', 'Gasto'],
                var_name='Tipo',
                value_name='Valor'
            )
            fig_orcamento = px.bar(
                df_plot,
                x='Categoria',
                y='Valor',
                color='Tipo',
                barmode='group',
                title='Comparativo Orçado vs. Gasto por Categoria',
                labels={'Valor': 'Valor (R$)', 'Categoria': 'Categoria'},
                text_auto='.2s'
            )
            fig_orcamento.update_traces(textposition='outside')
            fig_orcamento.update_layout(yaxis_tickformat="R$,.2f")
            st.plotly_chart(fig_orcamento, use_container_width=True)

            # 7. Tabela Detalhada
            st.markdown("### Tabela de Acompanhamento")
            df_display = df_analise[['Categoria', 'Percentual', 'Orcado', 'Gasto', 'Saldo', '%_Uso']]
            st.dataframe(
                df_display.style
                .format({
                    'Percentual': '{:.1f}%',
                    'Orcado': 'R$ {:,.2f}',
                    'Gasto': 'R$ {:,.2f}',
                    'Saldo': 'R$ {:,.2f}',
                    '%_Uso': '{:.1f}%'
                }),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.warning("A Renda Líquida para o período selecionado é zero ou negativa. O cálculo do orçamento não pode ser realizado.")

    # --- LÓGICA DOS MODAIS DE LANÇAMENTO ---

    # Função genérica para salvar dados no Excel
    def save_transaction(df_new, sheet_name):
        try:
            with st.spinner(f"Salvando em {sheet_name}..."):
                # Usar um lock para evitar concorrência no futuro
                from threading import Lock
                excel_lock = Lock()
                with excel_lock:
                    dfs = pd.read_excel(excel_path, sheet_name=None)
                    
                    sheet_df = dfs.get(sheet_name)
                    
                    # Garante a consistência dos tipos de dados antes de concatenar
                    for col in sheet_df.columns:
                        if col in df_new.columns:
                            if pd.api.types.is_datetime64_any_dtype(sheet_df[col]):
                                df_new[col] = pd.to_datetime(df_new[col])
                            else:
                                df_new[col] = df_new[col].astype(sheet_df[col].dtype, errors='ignore')
                    
                    updated_df = pd.concat([sheet_df, df_new], ignore_index=True)
                    dfs[sheet_name] = updated_df

                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        for name, df_sheet in dfs.items():
                            df_sheet.to_excel(writer, sheet_name=name, index=False)
            return True
        except Exception as e:
            st.error(f"Erro ao salvar no Excel: {e}")
            return False

    # Modal para Nova Despesa
    if st.session_state.get("show_despesa_form", False):
        st.subheader("📝 Lançar Nova Despesa")
        with st.container():
            # Ler dados para os selects
            df_conta = pd.read_excel(xls, sheet_name='Conta')
            contas = df_conta['Contas'].dropna().unique().tolist()
            df_cat_desp = pd.read_excel(xls, sheet_name='Despesas Categoria')
            categorias_desp = df_cat_desp.columns.tolist()

            # Categoria fica fora do form para reatividade
            categoria_despesa = st.selectbox("1. Selecione a Categoria", sorted(categorias_desp), key="cat_despesa")

            # Carregar itens da categoria selecionada dinamicamente
            itens_despesas, _ = carregar_itens_categoria()
            itens_categoria = itens_despesas.get(categoria_despesa, [])

            with st.form("despesa_form_reactive", clear_on_submit=True):
                st.info(f"Categoria selecionada: **{categoria_despesa}**")

                c1, c2 = st.columns(2)
                data_despesa = c1.date_input("Data", datetime.now(), key="data_despesa")
                valor_despesa = c2.number_input("Valor (negativo)", value=0.0, format="%.2f", key="valor_despesa")
                
                # Campo de descrição com lista de itens atualizada
                st.write("**2. Descrição**")
                opcao_descricao = st.selectbox(
                    "Selecione um item ou digite livremente:",
                    ["Digite livremente"] + sorted(itens_categoria),
                    key="opcao_desc_despesa"
                )
                
                if opcao_descricao == "Digite livremente":
                    descricao_despesa = st.text_input("Digite a nova descrição:", key="desc_despesa_livre")
                else:
                    descricao_despesa = opcao_descricao
                
                c3, c4 = st.columns(2)
                conta_despesa = c3.selectbox("3. Conta de Origem", sorted(contas), key="conta_despesa")
                pago_despesa = c4.selectbox("4. Status do Pagamento", ["Pago", "Pendente"], key="pago_despesa")

                favorecido_despesa = st.text_input("5. Favorecido", key="fav_despesa")
                
                # Converter para valor numérico (1 = Pago, 0 = Pendente)
                pago_valor = 1.0 if pago_despesa == "Pago" else 0.0

                col_submit, col_cancel = st.columns(2)
                submitted = col_submit.form_submit_button("✔️ Salvar Despesa", use_container_width=True, type="primary")
                if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                    st.session_state.show_despesa_form = False
                    st.rerun()

                if submitted:
                    if not descricao_despesa or valor_despesa == 0:
                        st.warning("Descrição e Valor são obrigatórios.")
                    elif valor_despesa > 0:
                        st.warning("O valor da despesa deve ser negativo.")
                    else:
                        new_data = pd.DataFrame([{
                            'DATA': pd.to_datetime(data_despesa), 
                            'FAVORECIDO': favorecido_despesa if favorecido_despesa else 'N/A',
                            'DESCRIÇÃO': descricao_despesa, 
                            'CATEGORIA': categoria_despesa, 
                            'CONTA': conta_despesa,
                            'FORMA DE PAGAMENTO': 'N/A', 
                            'VALOR': valor_despesa,
                            'PAGO': pago_valor
                        }])
                        if save_transaction(new_data, "Despesas"):
                            st.success("Despesa salva com sucesso!")
                            st.session_state.show_despesa_form = False
                            st.rerun()

    # Modal para Nova Receita
    if st.session_state.get("show_receita_form", False):
        st.subheader("💰 Lançar Nova Receita")
        with st.container():
            # Ler dados para os selects
            df_conta = pd.read_excel(xls, sheet_name='Conta')
            contas = df_conta['Contas'].dropna().unique().tolist()
            df_cat_rec = pd.read_excel(xls, sheet_name='Receitas Categoria')
            categorias_rec = df_cat_rec['SUBCATEGORIA'].dropna().unique().tolist()
            
            # Carregar itens de receitas (que não dependem da categoria)
            _, itens_receitas = carregar_itens_categoria()

            with st.form("receita_form_final", clear_on_submit=True):
                c1, c2 = st.columns(2)
                categoria_receita = c1.selectbox("Categoria", sorted(categorias_rec), key="cat_receita")
                conta_receita = c2.selectbox("Conta de Destino", sorted(contas), key="conta_receita")

                c3, c4 = st.columns(2)
                data_receita = c3.date_input("Data", datetime.now(), key="data_receita")
                valor_receita = c4.number_input("Valor (positivo)", value=0.00, format="%.2f", min_value=0.0, key="valor_receita")

                st.write("**Descrição**")
                opcao_descricao_rec = st.selectbox(
                    "Selecione um item ou digite livremente:",
                    ["Digite livremente"] + sorted(itens_receitas), # Usar a lista direta
                    key="opcao_desc_receita"
                )
                
                if opcao_descricao_rec == "Digite livremente":
                    descricao_receita = st.text_input("Digite a nova descrição:", key="desc_receita_livre")
                else:
                    descricao_receita = opcao_descricao_rec

                favorecido_receita = st.text_input("Pagador", key="fav_receita")

                col_submit, col_cancel = st.columns(2)
                submitted = col_submit.form_submit_button("✔️ Salvar Receita", use_container_width=True, type="primary")
                if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                    st.session_state.show_receita_form = False
                    st.rerun()

                if submitted:
                    if not descricao_receita or valor_receita == 0:
                        st.warning("Descrição e Valor são obrigatórios.")
                    else:
                        new_data = pd.DataFrame([{
                            'DATA': pd.to_datetime(data_receita), 
                            'CATEGORIA': categoria_receita,
                            'DESCRIÇÃO': descricao_receita, 
                            'CONTA': conta_receita, 
                            'VALOR': valor_receita,
                            'FAVORECIDO': favorecido_receita if favorecido_receita else 'N/A'
                        }])
                        if save_transaction(new_data, "Receitas"):
                            st.success("Receita salva com sucesso!")
                            st.session_state.show_receita_form = False
                            st.rerun()

    # Modal para Nova Venda
    if st.session_state.get("show_venda_form", False):
        st.subheader("🛒 Lançar Nova Venda")
        with st.container():
            # Carregar dados para os selects
            df_conta = pd.read_excel(xls, sheet_name='Conta')
            contas = df_conta['Contas'].dropna().unique().tolist()
            df_vendas_base = pd.read_excel(xls, sheet_name='Vendas')
            tipos_recebimento = df_vendas_base['TIPO DE RECEBIMENTO'].dropna().unique().tolist()
            clientes = df_vendas_base['Cliente'].dropna().unique().tolist()

            with st.form("venda_form", clear_on_submit=True):
                st.subheader("Preencha os dados da venda:")
                
                c1, c2 = st.columns(2)
                data_venda = c1.date_input("Data da Venda", datetime.now(), key="data_venda")
                valor_venda = c2.number_input("Valor da Venda", min_value=0.0, format="%.2f", key="valor_venda")

                c_desc, c_cli = st.columns(2)
                descricao_venda = c_desc.text_input("Descrição da Venda", key="desc_venda")
                
                # Modificação para sugestão de clientes
                opcoes_cliente = ["Adicionar novo cliente"] + sorted(clientes)
                cliente_selecionado = c_cli.selectbox("Cliente", opcoes_cliente, key="cliente_venda_select")
                
                if cliente_selecionado == "Adicionar novo cliente":
                    cliente_final = st.text_input("Nome do Novo Cliente", key="cliente_venda_novo")
                else:
                    cliente_final = cliente_selecionado


                c3, c4, c5 = st.columns(3)
                conta_venda = c3.selectbox("Conta de Destino", sorted(contas), key="conta_venda")
                tipo_receb_venda = c4.selectbox("Tipo de Recebimento", sorted(tipos_recebimento), key="tipo_receb_venda")
                status_venda = c5.selectbox("Status", ["Sim", "Não"], key="pago_venda")
                
                col_submit, col_cancel = st.columns(2)
                submitted = col_submit.form_submit_button("✔️ Salvar Venda", use_container_width=True, type="primary")
                if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                    st.session_state.show_venda_form = False
                    st.rerun()

                if submitted:
                    if not descricao_venda or valor_venda == 0:
                        st.warning("Descrição e Valor são obrigatórios.")
                    elif not cliente_final:
                        st.warning("O campo Cliente é obrigatório.")
                    else:
                        new_data = pd.DataFrame([{
                            'DATA': pd.to_datetime(data_venda),
                            'DESCRIÇÃO': descricao_venda,
                            'Cliente': cliente_final,
                            'CONTA': conta_venda,
                            'TIPO DE RECEBIMENTO': tipo_receb_venda,
                            'VALOR': valor_venda,
                            'Status': status_venda
                        }])
                        if save_transaction(new_data, "Vendas"):
                            st.success("Venda salva com sucesso!")
                            st.session_state.show_venda_form = False
                            st.rerun()

    # Modal para Nova Compra no Cartão
    if st.session_state.get("show_cc_form", False):
        st.subheader("💳 Lançar Nova Compra no Cartão")
        with st.container():
            with st.form("cc_form", clear_on_submit=True):
                st.subheader("Preencha os dados da compra:")
                df_cc_base = pd.read_excel(xls, sheet_name='Div_CC')
                cartoes = df_cc_base['Cartão'].dropna().unique().tolist()
                tipos_compra = df_cc_base['Tipo de Compra'].dropna().unique().tolist()

                c1, c2 = st.columns(2)
                data_cc = c1.date_input("Data da Compra", datetime.now(), key="data_cc")
                descricao_cc = c2.text_input("Descrição da Compra", key="desc_cc")
                
                c3, c4 = st.columns(2)
                tipo_compra_cc = c3.selectbox("Tipo de Compra", sorted(tipos_compra), key="tipo_compra_cc")
                cartao_cc = c4.selectbox("Cartão Utilizado", sorted(cartoes), key="cartao_cc")
                
                c5, c6, c7 = st.columns(3)
                qtd_parcelas = c5.number_input("Qtd. Parcelas", min_value=1, value=1, step=1, key="qtd_parcelas")
                valor_total_cc = c6.number_input("Valor Total", min_value=0.0, value=0.0, format="%.2f", key="valor_total_cc")
                valor_parcela = valor_total_cc / qtd_parcelas if qtd_parcelas > 0 else 0
                c7.metric("Valor da Parcela", f"{valor_parcela:.2f}")

                def calcular_vencimento(data_compra):
                    if data_compra.month == 12: return datetime(data_compra.year + 1, 1, 15)
                    return datetime(data_compra.year, data_compra.month + 1, 15)

                col_submit, col_cancel = st.columns(2)
                submitted = col_submit.form_submit_button("✔️ Salvar Compra", use_container_width=True, type="primary")
                if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                    st.session_state.show_cc_form = False
                    st.rerun()
                
                if submitted:
                    if not descricao_cc or valor_total_cc == 0:
                        st.warning("Descrição e Valor Total são obrigatórios.")
                    else:
                        new_data = pd.DataFrame([{
                            'Data': pd.to_datetime(data_cc), 'Descrição': descricao_cc, 'Tipo de Compra': tipo_compra_cc,
                            'Cartão': cartao_cc, 'Quantidade de parcelas': qtd_parcelas, 'Valor das parcelas': valor_parcela,
                            'valor total da compra': valor_total_cc, 'Situação': 'Pendente',
                            'Vencimento da Fatura': calcular_vencimento(data_cc)
                        }])
                        if save_transaction(new_data, "Div_CC"):
                            st.success("Compra salva com sucesso!")
                            st.session_state.show_cc_form = False
                            st.rerun()

    # --- FORMULÁRIOS DE EDIÇÃO ---
    
    # Formulário de edição de Vendas
    if st.session_state.get("show_edit_Vendas", False):
        st.subheader("✏️ Editar Venda")
        with st.container():
            # Carregar dados filtrados de vendas
            df_vendas = pd.read_excel(xls, sheet_name='Vendas')
            df_vendas['DATA'] = pd.to_datetime(df_vendas['DATA'], errors='coerce')
            df_vendas = df_vendas.dropna(subset=['DATA'])
            df_vendas['Ano'] = df_vendas['DATA'].dt.year.astype(str)
            df_vendas['Mês'] = df_vendas['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            vendas_filtradas = df_vendas[
                (df_vendas['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
                (df_vendas['Mês'].isin(meses_selecionados) if meses_selecionados else True)
            ]
            
            if not vendas_filtradas.empty:
                # Carregar dados para os selects
                df_conta = pd.read_excel(xls, sheet_name='Conta')
                contas = df_conta['Contas'].dropna().unique().tolist()
                tipos_recebimento = df_vendas['TIPO DE RECEBIMENTO'].dropna().unique().tolist()
                clientes = df_vendas['Cliente'].dropna().unique().tolist()
                
                # Criar opções para seleção
                opcoes_venda = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['Cliente']} - R$ {row['VALOR']:,.2f}" 
                               for idx, row in vendas_filtradas.iterrows()]
                
                venda_selecionada = st.selectbox("Selecionar venda para editar:", opcoes_venda, key="edit_venda_select")
                
                if venda_selecionada:
                    idx_selecionado = opcoes_venda.index(venda_selecionada)
                    row_to_edit = vendas_filtradas.iloc[idx_selecionado]
                    
                    with st.form("edit_venda_form"):
                        st.write("**Editar dados da venda:**")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            nova_data = col1.date_input("Data", value=pd.to_datetime(row_to_edit['DATA']).date(), key="edit_data_venda")
                            novo_cliente = col1.text_input("Cliente", value=row_to_edit['Cliente'], key="edit_cliente_venda")
                        with col2:
                            nova_descricao = col2.text_input("Descrição", value=row_to_edit['DESCRIÇÃO'], key="edit_descricao_venda")
                            novo_valor = col2.number_input("Valor", value=float(row_to_edit['VALOR']), format="%.2f", key="edit_valor_venda")
                        
                        col3, col4, col5 = st.columns(3)
                        with col3:
                            nova_conta = col3.selectbox("Conta", sorted(contas), index=contas.index(row_to_edit['CONTA']) if row_to_edit['CONTA'] in contas else 0, key="edit_conta_venda")
                        with col4:
                            novo_tipo_receb = col4.selectbox("Tipo de Recebimento", sorted(tipos_recebimento), index=tipos_recebimento.index(row_to_edit['TIPO DE RECEBIMENTO']) if row_to_edit['TIPO DE RECEBIMENTO'] in tipos_recebimento else 0, key="edit_tipo_receb_venda")
                        with col5:
                            novo_status = col5.selectbox("Status", ["Sim", "Não"], index=0 if row_to_edit['Status'] == "Sim" else 1, key="edit_status_venda")
                        
                        col_submit, col_cancel = st.columns(2)
                        submitted = col_submit.form_submit_button("💾 Salvar Alterações", use_container_width=True, type="primary")
                        if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                            st.session_state.show_edit_Vendas = False
                            st.rerun()
                        
                        if submitted:
                            updated_data = {
                                'DATA': pd.to_datetime(nova_data),
                                'Cliente': novo_cliente,
                                'DESCRIÇÃO': nova_descricao,
                                'CONTA': nova_conta,
                                'TIPO DE RECEBIMENTO': novo_tipo_receb,
                                'VALOR': novo_valor,
                                'Status': novo_status
                            }
                            
                            # Encontra o índice original no dataframe completo
                            original_idx = vendas_filtradas.index[idx_selecionado]
                            success, message = crud_system.update_record("Vendas", original_idx, updated_data)
                            
                            if success:
                                st.success("Venda atualizada com sucesso!")
                                st.session_state.show_edit_Vendas = False
                                st.rerun()
                            else:
                                st.error(f"Erro ao atualizar venda: {message}")
            else:
                st.info("Nenhuma venda encontrada com os filtros selecionados.")
    
    # Formulário de edição de Investimentos
    if st.session_state.get("show_edit_Investimentos", False):
        st.subheader("✏️ Editar Investimento")
        with st.container():
            # Carregar dados filtrados de investimentos
            df_investimentos = pd.read_excel(xls, sheet_name='Investimentos')
            df_investimentos['DATA'] = pd.to_datetime(df_investimentos['DATA'], errors='coerce')
            df_investimentos = df_investimentos.dropna(subset=['DATA'])
            df_investimentos['Ano'] = df_investimentos['DATA'].dt.year.astype(str)
            df_investimentos['Mês'] = df_investimentos['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            investimentos_filtrados = df_investimentos[
                (df_investimentos['Ano'].isin(st.session_state.get('ano_invest', [])) if st.session_state.get('ano_invest') else True) &
                (df_investimentos['Mês'].isin(st.session_state.get('mes_invest', [])) if st.session_state.get('mes_invest') else True) &
                ((df_investimentos['TIPO'] == st.session_state.get('tipo_invest', 'Todos')) | (st.session_state.get('tipo_invest', 'Todos') == 'Todos')) &
                ((df_investimentos['OBJETIVO'] == st.session_state.get('objetivo_invest', 'Todos')) | (st.session_state.get('objetivo_invest', 'Todos') == 'Todos')) &
                ((df_investimentos['ATIVO'] == st.session_state.get('ativo_invest', 'Todos')) | (st.session_state.get('ativo_invest', 'Todos') == 'Todos'))
            ]
            
            if not investimentos_filtrados.empty:
                # Criar opções para seleção
                opcoes_invest = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['ATIVO']} - R$ {row['VALOR_APORTE']:,.2f}" 
                                for idx, row in investimentos_filtrados.iterrows()]
                
                invest_selecionado = st.selectbox("Selecionar investimento para editar:", opcoes_invest, key="edit_invest_select")
                
                if invest_selecionado:
                    idx_selecionado = opcoes_invest.index(invest_selecionado)
                    row_to_edit = investimentos_filtrados.iloc[idx_selecionado]
                    
                    with st.form("edit_invest_form"):
                        st.write("**Editar dados do investimento:**")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            nova_data = col1.date_input("Data", value=pd.to_datetime(row_to_edit['DATA']).date(), key="edit_data_invest")
                            novo_tipo = col1.text_input("Tipo", value=row_to_edit['TIPO'], key="edit_tipo_invest")
                        with col2:
                            novo_ativo = col2.text_input("Ativo", value=row_to_edit['ATIVO'], key="edit_ativo_invest")
                            novo_valor = col2.number_input("Valor Aportado", value=float(row_to_edit['VALOR_APORTE']), format="%.2f", key="edit_valor_invest")
                        
                        col3, col4 = st.columns(2)
                        with col3:
                            nova_quantidade = col3.number_input("Quantidade", value=float(row_to_edit['QUANTIDADE']) if pd.notna(row_to_edit['QUANTIDADE']) else 0.0, format="%.2f", key="edit_quantidade_invest")
                            novo_preco_medio = col3.number_input("Preço Médio", value=float(row_to_edit['PRECO_MEDIO']) if pd.notna(row_to_edit['PRECO_MEDIO']) else 0.0, format="%.2f", key="edit_preco_medio_invest")
                        with col4:
                            novo_objetivo = col4.text_input("Objetivo", value=row_to_edit['OBJETIVO'], key="edit_objetivo_invest")
                        
                        col_submit, col_cancel = st.columns(2)
                        submitted = col_submit.form_submit_button("💾 Salvar Alterações", use_container_width=True, type="primary")
                        if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                            st.session_state.show_edit_Investimentos = False
                            st.rerun()
                        
                        if submitted:
                            updated_data = {
                                'DATA': pd.to_datetime(nova_data),
                                'TIPO': novo_tipo,
                                'ATIVO': novo_ativo,
                                'VALOR_APORTE': novo_valor,
                                'QUANTIDADE': nova_quantidade,
                                'PRECO_MEDIO': novo_preco_medio,
                                'OBJETIVO': novo_objetivo
                            }
                            
                            # Encontra o índice original no dataframe completo
                            original_idx = investimentos_filtrados.index[idx_selecionado]
                            success, message = crud_system.update_record("Investimentos", original_idx, updated_data)
                            
                            if success:
                                st.success("Investimento atualizado com sucesso!")
                                st.session_state.show_edit_Investimentos = False
                                st.rerun()
                            else:
                                st.error(f"Erro ao atualizar investimento: {message}")
            else:
                st.info("Nenhum investimento encontrado com os filtros selecionados.")
    
    # Formulário de edição de Cartão de Crédito
    if st.session_state.get("show_edit_Div_CC", False):
        st.subheader("✏️ Editar Compra no Cartão")
        with st.container():
            # Carregar dados filtrados de cartão de crédito
            df_cc = pd.read_excel(xls, sheet_name='Div_CC')
            df_cc['Data'] = pd.to_datetime(df_cc['Data'], errors='coerce')
            df_cc.dropna(subset=['Data'], inplace=True)
            df_cc['Ano'] = df_cc['Data'].dt.year.astype(str)
            df_cc['Mês'] = df_cc['Data'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            cc_filtrado = df_cc[
                (df_cc['Ano'].isin(st.session_state.ano_cc_sidebar) if st.session_state.ano_cc_sidebar else True) &
                (df_cc['Mês'].isin(st.session_state.mes_cc_sidebar) if st.session_state.mes_cc_sidebar else True) &
                ((df_cc['Cartão'] == st.session_state.cartao_cc_sidebar) | (st.session_state.cartao_cc_sidebar == 'Todos')) &
                ((df_cc['Situação'] == st.session_state.situacao_cc_sidebar) | (st.session_state.situacao_cc_sidebar == 'Todas')) &
                ((df_cc['Tipo de Compra'] == st.session_state.tipo_compra_cc_sidebar) | (st.session_state.tipo_compra_cc_sidebar == 'Todos')) &
                ((df_cc['Quantidade de parcelas'].astype(str) == st.session_state.parcelas_cc_sidebar) | (st.session_state.parcelas_cc_sidebar == 'Todas'))
            ]
            
            if not cc_filtrado.empty:
                # Criar opções para seleção
                opcoes_cc = [f"{row['Data'].strftime('%d/%m/%Y')} - {row['Descrição']} - R$ {row['valor total da compra']:,.2f}" 
                            for idx, row in cc_filtrado.iterrows()]
                
                cc_selecionado = st.selectbox("Selecionar compra para editar:", opcoes_cc, key="edit_cc_select")
                
                if cc_selecionado:
                    idx_selecionado = opcoes_cc.index(cc_selecionado)
                    row_to_edit = cc_filtrado.iloc[idx_selecionado]
                    
                    with st.form("edit_cc_form"):
                        st.write("**Editar dados da compra:**")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            nova_data = col1.date_input("Data", value=pd.to_datetime(row_to_edit['Data']).date(), key="edit_data_cc")
                            nova_descricao = col1.text_input("Descrição", value=row_to_edit['Descrição'], key="edit_descricao_cc")
                        with col2:
                            novo_tipo_compra = col2.text_input("Tipo de Compra", value=row_to_edit['Tipo de Compra'], key="edit_tipo_compra_cc")
                            novo_cartao = col2.text_input("Cartão", value=row_to_edit['Cartão'], key="edit_cartao_cc")
                        
                        col3, col4 = st.columns(2)
                        with col3:
                            nova_qtd_parcelas = col3.number_input("Qtd. Parcelas", value=int(row_to_edit['Quantidade de parcelas']), key="edit_qtd_parcelas_cc")
                            novo_valor_total = col3.number_input("Valor Total", value=float(row_to_edit['valor total da compra']), format="%.2f", key="edit_valor_total_cc")
                        with col4:
                            nova_situacao = col4.selectbox("Situação", ["Pendente", "Pago"], index=0 if row_to_edit['Situação'] == "Pendente" else 1, key="edit_situacao_cc")
                        
                        col_submit, col_cancel = st.columns(2)
                        submitted = col_submit.form_submit_button("💾 Salvar Alterações", use_container_width=True, type="primary")
                        if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                            st.session_state.show_edit_Div_CC = False
                            st.rerun()
                        
                        if submitted:
                            updated_data = {
                                'Data': pd.to_datetime(nova_data),
                                'Descrição': nova_descricao,
                                'Tipo de Compra': novo_tipo_compra,
                                'Cartão': novo_cartao,
                                'Quantidade de parcelas': nova_qtd_parcelas,
                                'valor total da compra': novo_valor_total,
                                'Situação': nova_situacao
                            }
                            
                            # Encontra o índice original no dataframe completo
                            original_idx = cc_filtrado.index[idx_selecionado]
                            success, message = crud_system.update_record("Div_CC", original_idx, updated_data)
                            
                            if success:
                                st.success("Compra atualizada com sucesso!")
                                st.session_state.show_edit_Div_CC = False
                                st.rerun()
                            else:
                                st.error(f"Erro ao atualizar compra: {message}")
            else:
                st.info("Nenhuma compra encontrada com os filtros selecionados.")

    # --- FORMULÁRIOS DE EXCLUSÃO ---
    
    # Formulário de exclusão de Vendas
    if st.session_state.get("show_delete_Vendas", False):
        st.subheader("🗑️ Excluir Venda")
        with st.container():
            # Carregar dados filtrados de vendas
            df_vendas = pd.read_excel(xls, sheet_name='Vendas')
            df_vendas['DATA'] = pd.to_datetime(df_vendas['DATA'], errors='coerce')
            df_vendas = df_vendas.dropna(subset=['DATA'])
            df_vendas['Ano'] = df_vendas['DATA'].dt.year.astype(str)
            df_vendas['Mês'] = df_vendas['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            vendas_filtradas = df_vendas[
                (df_vendas['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
                (df_vendas['Mês'].isin(meses_selecionados) if meses_selecionados else True)
            ]
            
            if not vendas_filtradas.empty:
                # Criar opções para seleção
                opcoes_venda = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['Cliente']} - R$ {row['VALOR']:,.2f}" 
                               for idx, row in vendas_filtradas.iterrows()]
                
                venda_selecionada = st.selectbox("Selecionar venda para excluir:", opcoes_venda, key="delete_venda_select")
                
                if st.button("🗑️ Confirmar Exclusão", key="confirm_delete_venda"):
                    if venda_selecionada:
                        idx_selecionado = opcoes_venda.index(venda_selecionada)
                        original_idx = vendas_filtradas.index[idx_selecionado]
                        success, message = crud_system.delete_record("Vendas", original_idx)
                        
                        if success:
                            st.success("Venda excluída com sucesso!")
                            st.session_state.show_delete_Vendas = False
                            st.rerun()
                        else:
                            st.error(f"Erro ao excluir venda: {message}")
            else:
                st.info("Nenhuma venda encontrada com os filtros selecionados.")
    
    # Formulário de exclusão de Investimentos
    if st.session_state.get("show_delete_Investimentos", False):
        st.subheader("🗑️ Excluir Investimento")
        with st.container():
            # Carregar dados filtrados de investimentos
            df_investimentos = pd.read_excel(xls, sheet_name='Investimentos')
            df_investimentos['DATA'] = pd.to_datetime(df_investimentos['DATA'], errors='coerce')
            df_investimentos = df_investimentos.dropna(subset=['DATA'])
            df_investimentos['Ano'] = df_investimentos['DATA'].dt.year.astype(str)
            df_investimentos['Mês'] = df_investimentos['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            investimentos_filtrados = df_investimentos[
                (df_investimentos['Ano'].isin(st.session_state.get('ano_invest', [])) if st.session_state.get('ano_invest') else True) &
                (df_investimentos['Mês'].isin(st.session_state.get('mes_invest', [])) if st.session_state.get('mes_invest') else True) &
                ((df_investimentos['TIPO'] == st.session_state.get('tipo_invest', 'Todos')) | (st.session_state.get('tipo_invest', 'Todos') == 'Todos')) &
                ((df_investimentos['OBJETIVO'] == st.session_state.get('objetivo_invest', 'Todos')) | (st.session_state.get('objetivo_invest', 'Todos') == 'Todos')) &
                ((df_investimentos['ATIVO'] == st.session_state.get('ativo_invest', 'Todos')) | (st.session_state.get('ativo_invest', 'Todos') == 'Todos'))
            ]
            
            if not investimentos_filtrados.empty:
                # Criar opções para seleção
                opcoes_invest = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['ATIVO']} - R$ {row['VALOR_APORTE']:,.2f}" 
                                for idx, row in investimentos_filtrados.iterrows()]
                
                invest_selecionado = st.selectbox("Selecionar investimento para excluir:", opcoes_invest, key="delete_invest_select")
                
                if st.button("🗑️ Confirmar Exclusão", key="confirm_delete_invest"):
                    if invest_selecionado:
                        idx_selecionado = opcoes_invest.index(invest_selecionado)
                        original_idx = investimentos_filtrados.index[idx_selecionado]
                        success, message = crud_system.delete_record("Investimentos", original_idx)
                        
                        if success:
                            st.success("Investimento excluído com sucesso!")
                            st.session_state.show_delete_Investimentos = False
                            st.rerun()
                        else:
                            st.error(f"Erro ao excluir investimento: {message}")
            else:
                st.info("Nenhum investimento encontrado com os filtros selecionados.")
    
    # Formulário de exclusão de Cartão de Crédito
    if st.session_state.get("show_delete_Div_CC", False):
        st.subheader("🗑️ Excluir Compra no Cartão")
        with st.container():
            # Carregar dados filtrados de cartão de crédito
            df_cc = pd.read_excel(xls, sheet_name='Div_CC')
            df_cc['Data'] = pd.to_datetime(df_cc['Data'], errors='coerce')
            df_cc.dropna(subset=['Data'], inplace=True)
            df_cc['Ano'] = df_cc['Data'].dt.year.astype(str)
            df_cc['Mês'] = df_cc['Data'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            cc_filtrado = df_cc[
                (df_cc['Ano'].isin(st.session_state.ano_cc_sidebar) if st.session_state.ano_cc_sidebar else True) &
                (df_cc['Mês'].isin(st.session_state.mes_cc_sidebar) if st.session_state.mes_cc_sidebar else True) &
                ((df_cc['Cartão'] == st.session_state.cartao_cc_sidebar) | (st.session_state.cartao_cc_sidebar == 'Todos')) &
                ((df_cc['Situação'] == st.session_state.situacao_cc_sidebar) | (st.session_state.situacao_cc_sidebar == 'Todas')) &
                ((df_cc['Tipo de Compra'] == st.session_state.tipo_compra_cc_sidebar) | (st.session_state.tipo_compra_cc_sidebar == 'Todos')) &
                ((df_cc['Quantidade de parcelas'].astype(str) == st.session_state.parcelas_cc_sidebar) | (st.session_state.parcelas_cc_sidebar == 'Todas'))
            ]
            
            if not cc_filtrado.empty:
                # Criar opções para seleção
                opcoes_cc = [f"{row['Data'].strftime('%d/%m/%Y')} - {row['Descrição']} - R$ {row['valor total da compra']:,.2f}" 
                            for idx, row in cc_filtrado.iterrows()]
                
                cc_selecionado = st.selectbox("Selecionar compra para excluir:", opcoes_cc, key="delete_cc_select")
                
                if st.button("🗑️ Confirmar Exclusão", key="confirm_delete_cc"):
                    if cc_selecionado:
                        idx_selecionado = opcoes_cc.index(cc_selecionado)
                        original_idx = cc_filtrado.index[idx_selecionado]
                        success, message = crud_system.delete_record("Div_CC", original_idx)
                        
                        if success:
                            st.success("Compra excluída com sucesso!")
                            st.session_state.show_delete_Div_CC = False
                            st.rerun()
                        else:
                            st.error(f"Erro ao excluir compra: {message}")
            else:
                st.info("Nenhuma compra encontrada com os filtros selecionados.")

    # --- FORMULÁRIOS DE EXCLUSÃO EM LOTE ---
    
    # Formulário de exclusão em lote de Vendas
    if st.session_state.get("show_bulk_delete_Vendas", False):
        st.subheader("🗑️ Exclusão em Lote - Vendas")
        with st.container():
            # Carregar dados filtrados de vendas
            df_vendas = pd.read_excel(xls, sheet_name='Vendas')
            df_vendas['DATA'] = pd.to_datetime(df_vendas['DATA'], errors='coerce')
            df_vendas = df_vendas.dropna(subset=['DATA'])
            df_vendas['Ano'] = df_vendas['DATA'].dt.year.astype(str)
            df_vendas['Mês'] = df_vendas['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            vendas_filtradas = df_vendas[
                (df_vendas['Ano'].isin(anos_selecionados) if anos_selecionados else True) &
                (df_vendas['Mês'].isin(meses_selecionados) if meses_selecionados else True)
            ]
            
            if not vendas_filtradas.empty:
                # Criar opções para seleção múltipla
                opcoes_venda = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['Cliente']} - R$ {row['VALOR']:,.2f}" 
                               for idx, row in vendas_filtradas.iterrows()]
                
                vendas_selecionadas = st.multiselect("Selecionar vendas para excluir:", opcoes_venda, key="bulk_delete_venda_select")
                
                if vendas_selecionadas:
                    st.warning(f"⚠️ Você está prestes a excluir {len(vendas_selecionadas)} venda(s). Esta ação não pode ser desfeita!")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("🗑️ Confirmar Exclusão em Lote", key="confirm_bulk_delete_venda", type="primary"):
                            success_count = 0
                            error_count = 0
                            
                            for venda_selecionada in vendas_selecionadas:
                                idx_selecionado = opcoes_venda.index(venda_selecionada)
                                original_idx = vendas_filtradas.index[idx_selecionado]
                                success, message = crud_system.delete_record("Vendas", original_idx)
                                
                                if success:
                                    success_count += 1
                                else:
                                    error_count += 1
                            
                            if success_count > 0:
                                st.success(f"✅ {success_count} venda(s) excluída(s) com sucesso!")
                            if error_count > 0:
                                st.error(f"❌ {error_count} venda(s) não puderam ser excluída(s).")
                            
                            st.session_state.show_bulk_delete_Vendas = False
                            st.rerun()
                    
                    with col2:
                        if st.button("✖️ Cancelar", key="cancel_bulk_delete_venda"):
                            st.session_state.show_bulk_delete_Vendas = False
                            st.rerun()
            else:
                st.info("Nenhuma venda encontrada com os filtros selecionados.")
    
    # Formulário de exclusão em lote de Investimentos
    if st.session_state.get("show_bulk_delete_Investimentos", False):
        st.subheader("🗑️ Exclusão em Lote - Investimentos")
        with st.container():
            # Carregar dados filtrados de investimentos
            df_investimentos = pd.read_excel(xls, sheet_name='Investimentos')
            df_investimentos['DATA'] = pd.to_datetime(df_investimentos['DATA'], errors='coerce')
            df_investimentos = df_investimentos.dropna(subset=['DATA'])
            df_investimentos['Ano'] = df_investimentos['DATA'].dt.year.astype(str)
            df_investimentos['Mês'] = df_investimentos['DATA'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            investimentos_filtrados = df_investimentos[
                (df_investimentos['Ano'].isin(st.session_state.get('ano_invest', [])) if st.session_state.get('ano_invest') else True) &
                (df_investimentos['Mês'].isin(st.session_state.get('mes_invest', [])) if st.session_state.get('mes_invest') else True) &
                ((df_investimentos['TIPO'] == st.session_state.get('tipo_invest', 'Todos')) | (st.session_state.get('tipo_invest', 'Todos') == 'Todos')) &
                ((df_investimentos['OBJETIVO'] == st.session_state.get('objetivo_invest', 'Todos')) | (st.session_state.get('objetivo_invest', 'Todos') == 'Todos')) &
                ((df_investimentos['ATIVO'] == st.session_state.get('ativo_invest', 'Todos')) | (st.session_state.get('ativo_invest', 'Todos') == 'Todos'))
            ]
            
            if not investimentos_filtrados.empty:
                # Criar opções para seleção múltipla
                opcoes_invest = [f"{row['DATA'].strftime('%d/%m/%Y')} - {row['ATIVO']} - R$ {row['VALOR_APORTE']:,.2f}" 
                                for idx, row in investimentos_filtrados.iterrows()]
                
                investimentos_selecionados = st.multiselect("Selecionar investimentos para excluir:", opcoes_invest, key="bulk_delete_invest_select")
                
                if investimentos_selecionados:
                    st.warning(f"⚠️ Você está prestes a excluir {len(investimentos_selecionados)} investimento(s). Esta ação não pode ser desfeita!")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("🗑️ Confirmar Exclusão em Lote", key="confirm_bulk_delete_invest", type="primary"):
                            success_count = 0
                            error_count = 0
                            
                            for invest_selecionado in investimentos_selecionados:
                                idx_selecionado = opcoes_invest.index(invest_selecionado)
                                original_idx = investimentos_filtrados.index[idx_selecionado]
                                success, message = crud_system.delete_record("Investimentos", original_idx)
                                
                                if success:
                                    success_count += 1
                                else:
                                    error_count += 1
                            
                            if success_count > 0:
                                st.success(f"✅ {success_count} investimento(s) excluído(s) com sucesso!")
                            if error_count > 0:
                                st.error(f"❌ {error_count} investimento(s) não puderam ser excluído(s).")
                            
                            st.session_state.show_bulk_delete_Investimentos = False
                            st.rerun()
                    
                    with col2:
                        if st.button("✖️ Cancelar", key="cancel_bulk_delete_invest"):
                            st.session_state.show_bulk_delete_Investimentos = False
                            st.rerun()
            else:
                st.info("Nenhum investimento encontrado com os filtros selecionados.")
    
    # Formulário de exclusão em lote de Cartão de Crédito
    if st.session_state.get("show_bulk_delete_Div_CC", False):
        st.subheader("🗑️ Exclusão em Lote - Compras no Cartão")
        with st.container():
            # Carregar dados filtrados de cartão de crédito
            df_cc = pd.read_excel(xls, sheet_name='Div_CC')
            df_cc['Data'] = pd.to_datetime(df_cc['Data'], errors='coerce')
            df_cc.dropna(subset=['Data'], inplace=True)
            df_cc['Ano'] = df_cc['Data'].dt.year.astype(str)
            df_cc['Mês'] = df_cc['Data'].dt.strftime('%b').str.capitalize().replace({'Feb': 'Fev', 'Apr': 'Abr', 'May': 'Mai', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Dec': 'Dez'})
            
            # Aplicar filtros
            cc_filtrado = df_cc[
                (df_cc['Ano'].isin(st.session_state.ano_cc_sidebar) if st.session_state.ano_cc_sidebar else True) &
                (df_cc['Mês'].isin(st.session_state.mes_cc_sidebar) if st.session_state.mes_cc_sidebar else True) &
                ((df_cc['Cartão'] == st.session_state.cartao_cc_sidebar) | (st.session_state.cartao_cc_sidebar == 'Todos')) &
                ((df_cc['Situação'] == st.session_state.situacao_cc_sidebar) | (st.session_state.situacao_cc_sidebar == 'Todas')) &
                ((df_cc['Tipo de Compra'] == st.session_state.tipo_compra_cc_sidebar) | (st.session_state.tipo_compra_cc_sidebar == 'Todos')) &
                ((df_cc['Quantidade de parcelas'].astype(str) == st.session_state.parcelas_cc_sidebar) | (st.session_state.parcelas_cc_sidebar == 'Todas'))
            ]
            
            if not cc_filtrado.empty:
                # Criar opções para seleção múltipla
                opcoes_cc = [f"{row['Data'].strftime('%d/%m/%Y')} - {row['Descrição']} - R$ {row['valor total da compra']:,.2f}" 
                            for idx, row in cc_filtrado.iterrows()]
                
                compras_selecionadas = st.multiselect("Selecionar compras para excluir:", opcoes_cc, key="bulk_delete_cc_select")
                
                if compras_selecionadas:
                    st.warning(f"⚠️ Você está prestes a excluir {len(compras_selecionadas)} compra(s). Esta ação não pode ser desfeita!")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("🗑️ Confirmar Exclusão em Lote", key="confirm_bulk_delete_cc", type="primary"):
                            success_count = 0
                            error_count = 0
                            
                            for compra_selecionada in compras_selecionadas:
                                idx_selecionado = opcoes_cc.index(compra_selecionada)
                                original_idx = cc_filtrado.index[idx_selecionado]
                                success, message = crud_system.delete_record("Div_CC", original_idx)
                                
                                if success:
                                    success_count += 1
                                else:
                                    error_count += 1
                            
                            if success_count > 0:
                                st.success(f"✅ {success_count} compra(s) excluída(s) com sucesso!")
                            if error_count > 0:
                                st.error(f"❌ {error_count} compra(s) não puderam ser excluída(s).")
                            
                            st.session_state.show_bulk_delete_Div_CC = False
                            st.rerun()
                    
                    with col2:
                        if st.button("✖️ Cancelar", key="cancel_bulk_delete_cc"):
                            st.session_state.show_bulk_delete_Div_CC = False
                            st.rerun()
            else:
                st.info("Nenhuma compra encontrada com os filtros selecionados.")

    # Modal para Novo Investimento
    if st.session_state.get("show_investimento_form", False):
        st.subheader("💰 Lançar Novo Investimento")
        with st.container():
            with st.form("investimento_form", clear_on_submit=True):
                st.subheader("Preencha os dados do investimento:")
                
                # Carregar dados existentes para sugestões
                df_investimentos_base = pd.read_excel(xls, sheet_name='Investimentos')
                tipos_investimento = df_investimentos_base['TIPO'].dropna().unique().tolist()
                ativos = df_investimentos_base['ATIVO'].dropna().unique().tolist()
                objetivos = df_investimentos_base['OBJETIVO'].dropna().unique().tolist()

                c1, c2 = st.columns(2)
                data_investimento = c1.date_input("Data do Investimento", datetime.now(), key="data_investimento")
                tipo_investimento = c2.selectbox("Tipo de Investimento", sorted(tipos_investimento), key="tipo_investimento")
                
                c3, c4 = st.columns(2)
                ativo = c3.text_input("Ativo", key="ativo_investimento")
                valor_aporte = c4.number_input("Valor Aportado", min_value=0.0, value=0.0, format="%.2f", key="valor_aporte")
                
                c5, c6 = st.columns(2)
                quantidade = c5.number_input("Quantidade", min_value=0.0, value=0.0, format="%.2f", key="quantidade_investimento")
                preco_medio = c6.number_input("Preço Médio", min_value=0.0, value=0.0, format="%.2f", key="preco_medio_investimento")
                
                objetivo = st.selectbox("Objetivo", sorted(objetivos), key="objetivo_investimento")

                col_submit, col_cancel = st.columns(2)
                submitted = col_submit.form_submit_button("✔️ Salvar Investimento", use_container_width=True, type="primary")
                if col_cancel.form_submit_button("✖️ Cancelar", use_container_width=True):
                    st.session_state.show_investimento_form = False
                    st.rerun()
                
                if submitted:
                    if not ativo or valor_aporte == 0:
                        st.warning("Ativo e Valor Aportado são obrigatórios.")
                    else:
                        new_data = pd.DataFrame([{
                            'DATA': pd.to_datetime(data_investimento),
                            'TIPO': tipo_investimento,
                            'ATIVO': ativo,
                            'VALOR_APORTE': valor_aporte,
                            'QUANTIDADE': quantidade,
                            'PRECO_MEDIO': preco_medio,
                            'OBJETIVO': objetivo
                        }])
                        if save_transaction(new_data, "Investimentos"):
                            st.success("Investimento salvo com sucesso!")
                            st.session_state.show_investimento_form = False
                            st.rerun()

except Exception as e:
    st.error("--- UM ERRO FATAL IMPEDIU O DASHBOARD DE CARREGAR ---")
    st.error("Isso geralmente é causado por um nome de aba incorreto no Excel ou um arquivo corrompido.")
    st.error(f"Detalhes do erro: {e}")
    import traceback
    st.code(traceback.format_exc())