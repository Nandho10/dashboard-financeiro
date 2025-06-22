import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
from backup_system import safe_backup
import numpy as np

class CRUDSystem:
    def __init__(self, excel_path='Base_financas.xlsx'):
        self.excel_path = excel_path
    
    def load_sheet_data(self, sheet_name):
        """Carrega dados de uma aba específica"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            return df
        except Exception as e:
            st.error(f"Erro ao carregar dados da aba {sheet_name}: {e}")
            return pd.DataFrame()
    
    def save_sheet_data(self, df, sheet_name):
        """Salva dados em uma aba específica"""
        try:
            # Cria backup antes de salvar
            safe_backup(f"before_{sheet_name}_update")
            
            # Carrega o workbook
            workbook = openpyxl.load_workbook(self.excel_path)
            
            # Remove a aba existente se existir
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
            
            # Cria nova aba
            worksheet = workbook.create_sheet(sheet_name)
            
            # Adiciona os dados
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Salva o workbook
            workbook.save(self.excel_path)
            workbook.close()
            
            return True, "Dados salvos com sucesso"
        except Exception as e:
            return False, f"Erro ao salvar dados: {e}"
    
    def update_record(self, sheet_name, row_index, updated_data):
        """Atualiza um registro específico"""
        try:
            df = self.load_sheet_data(sheet_name)
            if df.empty:
                return False, "Não foi possível carregar os dados"
            
            # Atualiza a linha específica
            for col, value in updated_data.items():
                if col in df.columns:
                    df.at[row_index, col] = value
            
            # Salva os dados atualizados
            success, message = self.save_sheet_data(df, sheet_name)
            if success:
                safe_backup(f"after_{sheet_name}_update")
            
            return success, message
        except Exception as e:
            return False, f"Erro ao atualizar registro: {e}"
    
    def delete_record(self, sheet_name, row_index):
        """Exclui um registro específico"""
        try:
            df = self.load_sheet_data(sheet_name)
            if df.empty:
                return False, "Não foi possível carregar os dados"
            
            # Remove a linha
            df = df.drop(index=row_index).reset_index(drop=True)
            
            # Salva os dados atualizados
            success, message = self.save_sheet_data(df, sheet_name)
            if success:
                safe_backup(f"after_{sheet_name}_delete")
            
            return success, message
        except Exception as e:
            return False, f"Erro ao excluir registro: {e}"
    
    def delete_multiple_records(self, sheet_name, row_indices):
        """Exclui múltiplos registros"""
        try:
            df = self.load_sheet_data(sheet_name)
            if df.empty:
                return False, "Não foi possível carregar os dados"
            
            # Remove as linhas
            df = df.drop(index=row_indices).reset_index(drop=True)
            
            # Salva os dados atualizados
            success, message = self.save_sheet_data(df, sheet_name)
            if success:
                safe_backup(f"after_{sheet_name}_bulk_delete")
            
            return success, message
        except Exception as e:
            return False, f"Erro ao excluir registros: {e}"

# Funções auxiliares para o dashboard
def format_dataframe_for_display(df, sheet_name):
    """Formata o dataframe para exibição no dashboard"""
    if df.empty:
        return df
    
    df_display = df.copy()
    
    # Formatação específica por aba
    if sheet_name == 'Despesas':
        if 'VALOR' in df_display.columns:
            df_display['VALOR'] = df_display['VALOR'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".") if pd.notna(x) else "R$ 0,00")
        if 'DATA' in df_display.columns:
            df_display['DATA'] = pd.to_datetime(df_display['DATA']).dt.strftime('%d/%m/%Y')
    
    elif sheet_name == 'Receitas':
        if 'VALOR' in df_display.columns:
            df_display['VALOR'] = df_display['VALOR'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".") if pd.notna(x) else "R$ 0,00")
        if 'DATA' in df_display.columns:
            df_display['DATA'] = pd.to_datetime(df_display['DATA']).dt.strftime('%d/%m/%Y')
    
    elif sheet_name == 'Vendas':
        if 'VALOR' in df_display.columns:
            df_display['VALOR'] = df_display['VALOR'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".") if pd.notna(x) else "R$ 0,00")
        if 'DATA' in df_display.columns:
            df_display['DATA'] = pd.to_datetime(df_display['DATA']).dt.strftime('%d/%m/%Y')
    
    elif sheet_name == 'Investimentos':
        if 'VALOR_APORTE' in df_display.columns:
            df_display['VALOR_APORTE'] = df_display['VALOR_APORTE'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".") if pd.notna(x) else "R$ 0,00")
        if 'DATA' in df_display.columns:
            df_display['DATA'] = pd.to_datetime(df_display['DATA']).dt.strftime('%d/%m/%Y')
    
    elif sheet_name == 'Div_CC':
        if 'valor total da compra' in df_display.columns:
            df_display['valor total da compra'] = df_display['valor total da compra'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".") if pd.notna(x) else "R$ 0,00")
        if 'Data' in df_display.columns:
            df_display['Data'] = pd.to_datetime(df_display['Data']).dt.strftime('%d/%m/%Y')
    
    return df_display

def create_editable_table(df, sheet_name, crud_system):
    """Cria uma tabela editável (sem botões de ação)"""
    if df.empty:
        st.info(f"Nenhum dado encontrado na aba {sheet_name}")
        return
    
    # Adiciona colunas de ação vazia para alinhamento, se necessário.
    # Pode ser removido se não houver mais botões na linha.
    # df['Ações'] = ''
    
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True
    ) 